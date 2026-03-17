# -*- coding: utf-8 -*-
"""
額分次數查詢系統（深色主題）- 更新版
- PageA：仁寶主名單 + 產出 GA09 喘息名單
- PageB：151（QP111）查詢
- PageC：仁寶 VS 151（需補分配）※此頁維持占位/可沿用你既有版本
- PageD：152（QP211）查「區間」→ 再用「精簡版架構」抓 QD120/QD120A（GA09 實際）
- PageE：151 名單（QP111輸出）vs QD120/QD120A（當月服務紀錄）之 BA 碼比對
        ✅ 新增：長照機構下拉選單（篩選 QD120/QD120A 僅計入相同機構的服務紀錄）
        ✅ 新規則：
            - 若 151 次數 >= QD120/QD120A 次數：沒問題（不呈現）
            - 若 151 次數 <  QD120/QD120A 次數：有問題（呈現並輸出）
"""

import re
import time
import threading
import calendar
import winreg
from datetime import datetime, timedelta

import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

from openpyxl import load_workbook
from openpyxl.styles import Font
import requests
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By


# =========================
# 顏色設定（深色主題）
# =========================
BG_MAIN   = "#121212"
BG_CARD   = "#1E1E1E"
BG_BTN    = "#2A2A2A"
BG_TEXT   = "#181818"
FG_MAIN   = "#FFFFFF"
FG_HINT   = "#9E9E9E"
FG_LINK   = "#64B5F6"
BORDER    = "#333333"

ACCENT_OK  = "#10B981"
ACCENT_GO  = "#3949AB"
ACCENT_SEL = "#2563EB"

# =========================
# 長照機構（PageE 下拉選單）
# =========================
ORG_NO_FILTER = "（不過濾）"
ORG_OPTIONS = [
    ORG_NO_FILTER,
    "大安心喜樂福祉有限公司私立大安心居家長照機構",
    "大慶健康事業股份有限公司私立大慶居家長照機構",
    "大慶健康事業股份有限公司私立誠善居家長照機構",
    "大慶健康事業股份有限公司私立誠欣居家長照機構",
    "摯誠健康事業股份有限公司私立欣源美居家長照機構"
]

def _norm_text(s: str) -> str:
    return str(s or "").replace(" ", "").replace("　", "").strip()

def get_chrome_major_version() -> int | None:
    registry_paths = [
        r"SOFTWARE\Google\Chrome\BLBeacon",
        r"SOFTWARE\WOW6432Node\Google\Chrome\BLBeacon",
    ]
    for root in (winreg.HKEY_CURRENT_USER, winreg.HKEY_LOCAL_MACHINE):
        for path in registry_paths:
            try:
                with winreg.OpenKey(root, path) as key:
                    version, _ = winreg.QueryValueEx(key, "version")
                    if version:
                        major = str(version).split(".")[0]
                        if major.isdigit():
                            return int(major)
            except Exception:
                continue
    return None

def _row_get_first_str(row: dict, keys: list[str]) -> str:
    for k in keys:
        v = row.get(k)
        if v is None:
            continue
        if isinstance(v, (dict, list, tuple)):
            continue
        s = str(v).strip()
        if s:
            return s
    return ""

def extract_title(row: dict) -> str:
    for k in ("title", "qd120Title", "serviceTitle", "itemTitle"):
        v = row.get(k)
        if v:
            return str(v).strip()
    return ""

def extract_org_from_row(row: dict) -> str:
    """
    從 QD120/QD120A row 盡可能抓出「長照機構/單位名稱」字串。
    LCMS 的 JSON 欄位在不同頁/權限下可能不同，因此做多鍵容錯。
    """
    candidates = [
        "pi400aName", "pi400Name", "pi400AName", "pi400", "pi400a",
        "qd120Pi400", "qd120Pi400Name", "qd120Pi400aName",
        "qd120APi400", "qd120APi400Name", "qd120APi400aName",
        "orgName", "organizationName", "unitName", "providerName",
        "servOrgName", "servUnitName",
    ]

    s = _row_get_first_str(row, candidates)
    if s:
        return s

    for k in ["pi400a", "pi400", "pi400Obj", "org", "provider"]:
        v = row.get(k)
        if isinstance(v, dict):
            for kk in ["name", "title", "pi400aName", "pi400Name"]:
                vv = v.get(kk)
                if vv:
                    return str(vv).strip()

    title = extract_title(row)
    return title

def row_matches_org(row: dict, org_filter: str | None) -> bool:
    """
    org_filter:
      - None / "" / ORG_NO_FILTER：不過濾
      - 其他：只保留「能對應到相同機構」的 row
    """
    if not org_filter or _norm_text(org_filter) == _norm_text(ORG_NO_FILTER):
        return True

    row_org = extract_org_from_row(row)
    if not row_org:
        return False

    return _norm_text(row_org) == _norm_text(org_filter)


# =========================
# LCMS endpoints
# =========================
LCMS_URL = "https://csms.mohw.gov.tw/lcms/"
API_QP111_FILTER = "https://csms.mohw.gov.tw/lcms/qp/filterQp111/"
API_QP211_FILTER = "https://csms.mohw.gov.tw/lcms/qp/filterQp211/"  # 152（喘息/區間）

EX_ID_URL = (
    "https://csms.mohw.gov.tw/lcms/ca/filter/"
    "?caseno=&serialno=&name=&idno=&pi400aName="
    "&_qd111Resp=&_qd1115Resp=&_qd111adjHint=&_qd111pi400Hint="
    "&_checkAa10Rsp=&_qp110adjHint=&_flexErrHint=&_ca110Modify=&_caseDeath="
    "&_uploadCc01=&_qp300HintPi400=&_qd300HintPi400=&_qd310HintPi400="
    "&_cmsLowerB=&_fh410NotExistsB="
    "&birthDt1=&birthDt2=&applyDt1=&applyDt2=&openDt1=&openDt2=&closeDt1=&closeDt2="
    "&qevalDt1=&qevalDt2=&qmaxInstructDt1=&qmaxInstructDt2=&processDt1=&processD2="
    "&applySource=&_hpCreated=&sextype=&applyType=&censusDiff=&aborigine=&raceType="
    "&liveType=&isdisbook=&sptype=&cmsLev=&disLev=&levcode=&isSick=&discode=&qcntcode="
    "&twnspcode=&vilgcode=&qinformCntcode=&informTwnspcode=&informVilgcode="
    "&qQd120ServDt_b=&qQd120ServDt_e=&sb210id=&sb500id=&ca113exists=&qca113title="
    "&qservUser=&sb400id=&doQuery=true&qdList1=yes&limit=100&offset=0&order=asc"
)

QD120_URL_TEMPLATE = (
    "https://csms.mohw.gov.tw/lcms/qd/filterQd120/{case_id}"
    "?doQuery=yes&ca100id={case_id}&perms=true"
    "&stype=&sourceType=&status="
    "&servDt1={servDt1}&servDt2={servDt2}"
    "&qd120Pi400=&servUserName=&aa10Status="
    "&order=asc&offset={offset}&limit={limit}"
)

QD120A_URL_TEMPLATE = (
    "https://csms.mohw.gov.tw/lcms/qd/filterQd120A/{case_id}"
    "?doQuery=yes&ca100id={case_id}&perms=true"
    "&stype=&sourceType=&status="
    "&servDt1={servDt1}&servDt2={servDt2}"
    "&qd120APi400=&servUserName=&servUserIdno=&aa10Status="
    "&order=asc&offset={offset}&limit={limit}"
)

REQ_HEADERS = {
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": "https://csms.mohw.gov.tw/lcms/",
}


def http_get_json_with_retry(
    sess: requests.Session,
    url: str,
    params: dict | None = None,
    headers: dict | None = None,
    timeout: int = 30,
    max_retries: int = 3,
) -> dict:
    """
    共用 GET(JSON) 請求，連線類錯誤最多重試 max_retries 次。
    """
    last_err = None
    retry_log_hook = getattr(sess, "_retry_log_hook", None)

    for attempt in range(1, max_retries + 1):
        try:
            resp = sess.get(url, params=params, headers=headers, timeout=timeout)
            resp.raise_for_status()
            return resp.json()
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
            last_err = e
            if attempt >= max_retries:
                break
            if callable(retry_log_hook):
                try:
                    retry_log_hook(f"⚠ 連線異常，進行重試 {attempt + 1}/{max_retries} ...")
                except Exception:
                    pass
            time.sleep(0.8 * attempt)
        except requests.exceptions.RequestException:
            raise
        except ValueError:
            raise

    raise requests.exceptions.ConnectionError(
        f"連線失敗（已重試{max_retries}次）：{last_err}"
    )


# =========================================================
# ttk 深色風格
# =========================================================
def setup_ttk_style(root: tk.Tk):
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    style.configure(
        "Treeview",
        background="#0F0F0F",
        fieldbackground="#0F0F0F",
        foreground=FG_MAIN,
        rowheight=26,
        bordercolor=BORDER,
        lightcolor=BORDER,
        darkcolor=BORDER
    )
    style.configure(
        "Treeview.Heading",
        background="#242424",
        foreground=FG_MAIN,
        relief="flat",
        font=("Microsoft JhengHei UI", 10, "bold")
    )
    style.map(
        "Treeview",
        background=[("selected", ACCENT_SEL)],
        foreground=[("selected", FG_MAIN)]
    )

    style.configure(
        "TProgressbar",
        troughcolor="#2B2B2B",
        background=ACCENT_OK,
        bordercolor="#2B2B2B",
        lightcolor=ACCENT_OK,
        darkcolor=ACCENT_OK
    )

    style.configure(
        "TCombobox",
        fieldbackground="#2D2D2D",
        background="#2D2D2D",
        foreground=FG_MAIN
    )


# =========================================================
# UI Blocking Dialog Helpers（可被背景執行緒呼叫）
# =========================================================
def ui_blocking_message(master: tk.Tk, title: str, message: str) -> bool:
    done = threading.Event()
    result = {"ok": False}

    def _show():
        ok = messagebox.askokcancel(title, message, parent=master)
        result["ok"] = bool(ok)
        done.set()

    master.after(0, _show)
    done.wait()
    return result["ok"]


def ui_blocking_input(master: tk.Tk, title: str, message: str, default: str = "", password: bool = False) -> str | None:
    done = threading.Event()
    result = {"value": None}

    def _show():
        win = tk.Toplevel(master)
        win.title(title)
        win.configure(bg=BG_CARD)
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text=message, bg=BG_CARD, fg=FG_MAIN,
                 font=("Microsoft JhengHei UI", 10)).pack(padx=16, pady=(14, 6), anchor="w")

        var = tk.StringVar(value=default)
        ent = tk.Entry(win, textvariable=var, width=28,
                       bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN,
                       relief="flat", show="*" if password else "")
        ent.pack(padx=16, pady=(0, 10), fill="x")
        ent.focus_set()

        btn_frame = tk.Frame(win, bg=BG_CARD)
        btn_frame.pack(padx=16, pady=(0, 14), fill="x")

        def ok():
            result["value"] = var.get()
            win.destroy()
            done.set()

        def cancel():
            result["value"] = None
            win.destroy()
            done.set()

        tk.Button(btn_frame, text="OK", command=ok,
                  bg=ACCENT_OK, fg="white", relief="flat", cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 6))
        tk.Button(btn_frame, text="取消", command=cancel,
                  bg="#4B5563", fg="white", relief="flat", cursor="hand2").pack(side="left", expand=True, fill="x", padx=(6, 0))

        win.protocol("WM_DELETE_WINDOW", cancel)

    master.after(0, _show)
    done.wait()
    return result["value"]


def ui_blocking_select_option(master: tk.Tk, title: str, message: str, options: list[str], default: str | None = None) -> str | None:
    done = threading.Event()
    result = {"value": None}

    def _show():
        win = tk.Toplevel(master)
        win.title(title)
        win.configure(bg=BG_CARD)
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text=message, bg=BG_CARD, fg=FG_MAIN,
                 font=("Microsoft JhengHei UI", 10)).pack(padx=16, pady=(14, 6), anchor="w")

        init_val = default if default in (options or []) else ((options or [""])[0])
        var = tk.StringVar(value=init_val)
        cmb = ttk.Combobox(win, textvariable=var, values=options, state="readonly", width=56)
        cmb.pack(padx=16, pady=(0, 10), fill="x")
        cmb.focus_set()

        btn_frame = tk.Frame(win, bg=BG_CARD)
        btn_frame.pack(padx=16, pady=(0, 14), fill="x")

        def ok():
            result["value"] = str(var.get()).strip()
            win.destroy()
            done.set()

        def cancel():
            result["value"] = None
            win.destroy()
            done.set()

        tk.Button(btn_frame, text="確定", command=ok,
                  bg=ACCENT_OK, fg="white", relief="flat", cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 6))
        tk.Button(btn_frame, text="取消", command=cancel,
                  bg="#4B5563", fg="white", relief="flat", cursor="hand2").pack(side="left", expand=True, fill="x", padx=(6, 0))

        win.protocol("WM_DELETE_WINDOW", cancel)

    master.after(0, _show)
    done.wait()
    return result["value"]


def ui_blocking_select_option(master: tk.Tk, title: str, message: str, options: list[str], default: str | None = None) -> str | None:
    done = threading.Event()
    result = {"value": None}

    def _show():
        win = tk.Toplevel(master)
        win.title(title)
        win.configure(bg=BG_CARD)
        win.resizable(False, False)
        win.grab_set()

        tk.Label(win, text=message, bg=BG_CARD, fg=FG_MAIN,
                 font=("Microsoft JhengHei UI", 10)).pack(padx=16, pady=(14, 6), anchor="w")

        init = default if (default in (options or [])) else ((options or [""])[0])
        var = tk.StringVar(value=init)
        cmb = ttk.Combobox(win, textvariable=var, values=options, state="readonly", width=56)
        cmb.pack(padx=16, pady=(0, 10), fill="x")
        cmb.focus_set()

        btn_frame = tk.Frame(win, bg=BG_CARD)
        btn_frame.pack(padx=16, pady=(0, 14), fill="x")

        def ok():
            result["value"] = str(var.get()).strip()
            win.destroy()
            done.set()

        def cancel():
            result["value"] = None
            win.destroy()
            done.set()

        tk.Button(btn_frame, text="確定", command=ok,
                  bg=ACCENT_OK, fg="white", relief="flat", cursor="hand2").pack(side="left", expand=True, fill="x", padx=(0, 6))
        tk.Button(btn_frame, text="取消", command=cancel,
                  bg="#4B5563", fg="white", relief="flat", cursor="hand2").pack(side="left", expand=True, fill="x", padx=(6, 0))

        win.protocol("WM_DELETE_WINDOW", cancel)

    master.after(0, _show)
    done.wait()
    return result["value"]


# =========================================================
# Selenium 登入 + Cookie → requests.Session
# =========================================================
driver = None
session = None


def is_cloudflare_page(driver):
    try:
        html = driver.page_source.lower()
        return ("cloudflare" in html and
                ("checking your browser" in html or
                 "access denied" in html or
                 "請啟用 javascript" in html))
    except Exception:
        return False


def wait_for_login_ready(driver, timeout=30):
    end = time.time() + timeout
    while time.time() < end:
        try:
            driver.find_element(By.ID, "username")
            driver.find_element(By.NAME, "password")
            return True
        except Exception:
            time.sleep(0.3)
    return False


def detect_recaptcha(driver):
    try:
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for iframe in iframes:
            src = iframe.get_attribute("src") or ""
            if "recaptcha" in src or "google.com/recaptcha" in src:
                return True
        return False
    except Exception:
        return False


def lcms_login_gui(master: tk.Tk, username: str, password: str) -> requests.Session | None:
    """
    UI 互動版本：
    - Cloudflare / reCAPTCHA：使用 messagebox 等待使用者完成
    - 圖形驗證碼：使用程式內輸入框（可留空）
    """
    global session, driver

    options = uc.ChromeOptions()
    options.add_argument("--no-sandbox")

    chrome_major = get_chrome_major_version()
    if chrome_major:
        driver = uc.Chrome(options=options, version_main=chrome_major)
    else:
        driver = uc.Chrome(options=options)
    driver.get(LCMS_URL)
    time.sleep(2)

    if is_cloudflare_page(driver):
        ok = ui_blocking_message(
            master,
            "需要人工驗證（Cloudflare）",
            "偵測到 Cloudflare 驗證。\n\n"
            "請在開啟的 Chrome 視窗完成驗證，完成後按「OK」繼續。\n"
            "若要取消，按「取消」。"
        )
        if not ok:
            try:
                driver.quit()
            except Exception:
                pass
            return None
        time.sleep(1)

    if not wait_for_login_ready(driver):
        ui_blocking_message(master, "登入失敗", "登入頁面載入逾時（找不到帳號密碼欄位）。")
        try:
            driver.quit()
        except Exception:
            pass
        return None

    # 填帳密
    try:
        acc = driver.find_element(By.ID, "username")
        pwd = driver.find_element(By.NAME, "password")
        acc.clear()
        acc.send_keys(username)
        pwd.clear()
        pwd.send_keys(password)
    except Exception:
        ui_blocking_message(master, "登入失敗", "無法填入帳號密碼，請確認頁面是否正常。")
        try:
            driver.quit()
        except Exception:
            pass
        return None

    # reCAPTCHA
    if detect_recaptcha(driver):
        ok = ui_blocking_message(
            master,
            "需要人工驗證（reCAPTCHA）",
            "偵測到 reCAPTCHA。\n\n"
            "請在 Chrome 視窗完成勾選/驗證後按「OK」繼續。\n"
            "若要取消，按「取消」。"
        )
        if not ok:
            try:
                driver.quit()
            except Exception:
                pass
            return None
        time.sleep(1)

    # 圖形驗證碼（若有欄位）→ UI 輸入
    try:
        captcha_value = ui_blocking_input(master, "圖形驗證碼", "如登入頁有圖形驗證碼，請輸入（沒有可留空）：", default="")
        if captcha_value is None:
            try:
                driver.quit()
            except Exception:
                pass
            return None
        captcha_value = captcha_value.strip()
        if captcha_value:
            driver.find_element(By.ID, "captcha").send_keys(captcha_value)
    except Exception:
        pass

    # 按登入（沿用你原本 XPath）
    try:
        login_btn = driver.find_element(
            By.XPATH,
            "//*[@id='id_content']/div/div/form/div[6]/div/div/input[1]"
        )
        login_btn.click()
        time.sleep(2)
    except Exception:
        ui_blocking_message(master, "登入失敗", "找不到登入按鈕（頁面結構可能變更）。")
        try:
            driver.quit()
        except Exception:
            pass
        return None

    # 取 cookies
    try:
        cookies = driver.get_cookies()
        if not cookies:
            ui_blocking_message(master, "登入失敗", "登入後無法取得 Cookie（可能未登入成功）。")
            try:
                driver.quit()
            except Exception:
                pass
            return None
    except Exception:
        ui_blocking_message(master, "登入失敗", "取得 Cookie 失敗。")
        try:
            driver.quit()
        except Exception:
            pass
        return None

    session = requests.Session()
    for c in cookies:
        session.cookies.set(c["name"], c["value"])

    try:
        driver.quit()
    except Exception:
        pass
    return session


# =========================================================
# 共用解析/規則（151/152）
# =========================================================
PAIR_RE = re.compile(r"([A-Z]{2}[0-9A-Za-z\-]+)\s*[:：]\s*(\d+)", re.IGNORECASE)

def parse_qd_pairs(text: str) -> dict:
    if not text:
        return {}
    found = PAIR_RE.findall(text)
    return {str(k).upper(): int(v) for k, v in found}

def roc_date_to_ym(last_updated: str) -> str | None:
    """
    114/12/29 -> 11412
    """
    s = (last_updated or "").strip()
    if not s:
        return None
    m = re.match(r"^\s*(\d{3})\s*/\s*(\d{1,2})", s)
    if not m:
        return None
    yyy = m.group(1)
    mm = str(int(m.group(2))).zfill(2)
    return f"{yyy}{mm}"

def parse_yyymm_range(range_str: str) -> tuple[int, int] | None:
    """
    "11501~11512" -> (11501, 11512)
    "11503" -> (11503, 11503)
    """
    s = (range_str or "").strip()
    if not s:
        return None
    if "~" in s:
        a, b = s.split("~", 1)
        a = a.strip()
        b = b.strip()
        if a.isdigit() and b.isdigit():
            return (int(a), int(b))
        return None
    if s.isdigit():
        v = int(s)
        return (v, v)
    return None

def ym_in_range(target_ym: str, range_str: str) -> bool:
    if not target_ym or (not target_ym.isdigit()):
        return False
    pr = parse_yyymm_range(range_str)
    if not pr:
        return False
    a, b = pr
    t = int(target_ym)
    return a <= t <= b


# =========================================================
# PageD/E 需要的：姓名→case_id→QD120/QD120A（民國日期）
# =========================================================
def Hex16(name: str) -> str:
    b = (name or "").encode("utf-8")
    hx = b.hex().upper()
    return "%" + "%".join(re.findall(".{2}", hx))

def roc_yyyymm_range_to_servDt_enc(roc_start_yyyymm: int, roc_end_yyyymm: int) -> tuple[str, str, str]:
    """
    11411~11510 -> servDt1=114%2F11%2F01 , servDt2=115%2F10%2F31
    回傳：(servDt1_enc, servDt2_enc, human_roc)
    """
    rs = str(roc_start_yyyymm).strip()
    re_ = str(roc_end_yyyymm).strip()
    if not (rs.isdigit() and re_.isdigit() and len(rs) == 5 and len(re_) == 5):
        raise ValueError(f"ROC 區間格式錯誤：{roc_start_yyyymm}~{roc_end_yyyymm}")

    y1 = int(rs[:3])
    m1 = int(rs[3:5])
    y2 = int(re_[:3])
    m2 = int(re_[3:5])
    if not (1 <= m1 <= 12 and 1 <= m2 <= 12):
        raise ValueError(f"月份需 01~12：{roc_start_yyyymm}~{roc_end_yyyymm}")

    g_y2 = y2 + 1911
    last_day = calendar.monthrange(g_y2, m2)[1]

    servDt1_enc = f"{y1}%2F{m1:02d}%2F01"
    servDt2_enc = f"{y2}%2F{m2:02d}%2F{last_day:02d}"
    human = f"{y1}/{m1:02d}/01 ~ {y2}/{m2:02d}/{last_day:02d}"
    return servDt1_enc, servDt2_enc, human

def roc_yyyymm_to_servDt_enc(roc_yyyymm: int) -> tuple[str, str, str]:
    """
    單月：11412 -> 114/12/01 ~ 114/12/31（encoded）
    """
    return roc_yyyymm_range_to_servDt_enc(roc_yyyymm, roc_yyyymm)

def api_case_find_case_ids_strict(sess: requests.Session, person: str) -> tuple[list[str], str]:
    """
    用 CA filter 找 case_id（優先完全同名）
    回傳：(case_ids, note)
    """
    name = (person or "").strip()
    if not name:
        return [], "姓名空白"

    url = EX_ID_URL.replace("name=", "name=" + Hex16(name))
    try:
        js = http_get_json_with_retry(
            sess,
            url,
            headers=REQ_HEADERS,
            timeout=30,
            max_retries=3,
        )
    except Exception as e:
        return [], f"CA 查詢失敗：{e}"

    rows = js.get("rows", []) or []
    if not rows:
        return [], "CA 查無 rows"

    exact_rows = [r for r in rows if str(r.get("name", "")).strip() == name]
    if exact_rows:
        ids = []
        for r in exact_rows:
            cid = r.get("id")
            if cid is not None and str(cid).strip():
                s = str(cid).strip()
                if s not in ids:
                    ids.append(s)
        if len(ids) > 1:
            return ids, f"完全同名出現多案（將加總）：{len(ids)} 案"
        return ids, ""

    if len(rows) == 1:
        cid = rows[0].get("id")
        if cid is None or (not str(cid).strip()):
            return [], "CA 回傳缺少 id"
        return [str(cid).strip()], "非完全同名但僅 1 筆，已採用該案（建議確認姓名完整性）"

    preview = []
    for r in rows[:5]:
        preview.append(f"{r.get('name','')}|id={r.get('id')}")
    return [], "同名/相似多筆，為避免抓錯已略過：" + "、".join(preview)

def qd_fetch_all_pages(sess: requests.Session, template: str, case_id: str, servDt1_enc: str, servDt2_enc: str,
                       limit: int = 100, max_pages: int = 80) -> tuple[list[dict], str]:
    rows_all: list[dict] = []
    offset = 0
    note = ""
    for _ in range(max_pages):
        url = template.format(case_id=case_id, servDt1=servDt1_enc, servDt2=servDt2_enc, offset=offset, limit=limit)
        try:
            js = http_get_json_with_retry(
                sess,
                url,
                headers=REQ_HEADERS,
                timeout=30,
                max_retries=3,
            )
        except Exception as e:
            return rows_all, f"QD 讀取失敗：{e}"

        rows = js.get("rows", []) or []
        if not rows:
            break
        rows_all.extend(rows)
        if len(rows) < limit:
            break
        offset += limit
        if offset > 50000:
            note = "QD 分頁 offset 過大，已中止"
            break
    return rows_all, note

def extract_code_from_title(title: str) -> str:
    """
    支援 BA16-1 這類含連字號碼別
    """
    t = (title or "").strip().upper()
    m = re.match(r"([A-Z]{2}[0-9A-Z\-]+)", t)
    return m.group(1) if m else ""

def extract_code_any(row: dict) -> str:
    """
    盡可能從 row 內找到服務碼（優先 title）
    """
    title = extract_title(row)
    code = extract_code_from_title(title)
    if code:
        return code

    for k in ("servCode", "serviceCode", "code", "qdCode", "qd100", "qd100s"):
        v = row.get(k)
        if not v:
            continue
        s = str(v).strip().upper()
        m = re.match(r"([A-Z]{2}[0-9A-Z\-]+)", s)
        if m:
            return m.group(1)
    return ""

def _safe_int(v, default=1) -> int:
    try:
        if v is None:
            return default
        s = str(v).strip()
        if s == "":
            return default
        n = int(float(s))
        return n if n > 0 else default
    except Exception:
        return default

def get_serv_count_guess(row: dict) -> int:
    """
    若系統 row 有次數欄位就拿；否則視為 1
    """
    for k in ("servCnt", "cnt", "times", "servTimes", "qty", "q", "num", "amount", "qdCnt"):
        if k in row:
            return _safe_int(row.get(k), default=1)
    return 1

def dedup_rows(rows: list[dict]) -> list[dict]:
    """
    盡量用 id 去重；沒有 id 就用日期/時間/title/服務員/狀態做組合去重
    """
    seen = set()
    out = []
    for r in rows:
        rid = r.get("id") or r.get("qd120Id") or r.get("qd120AId") or r.get("qd120id")
        if rid is not None and str(rid).strip():
            key = ("ID", str(rid).strip())
        else:
            key = (
                "CMP",
                str(r.get("servDt", "")).strip(),
                str(r.get("hhmm", "")).strip(),
                extract_title(r),
                str(r.get("servUserName", "")).strip() or str(r.get("servUser", "")).strip(),
                str(r.get("aa10Status", "")).strip(),
            )
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out

def merge_int_maps(a: dict, b: dict) -> dict:
    out = dict(a or {})
    for k, v in (b or {}).items():
        out[k] = out.get(k, 0) + int(v)
    return out

def format_ba_map(m: dict) -> str:
    if not m:
        return ""
    parts = [f"{k}:{m[k]}" for k in sorted(m.keys())]
    return "、".join(parts)

def compare_ba_alloc_vs_used(alloc: dict, used: dict) -> tuple[bool, str]:
    """
    新規則：
    - 151(alloc) >= QD(used) : 沒問題
    - 151(alloc) <  QD(used) : 有問題（回傳差異）
    回傳：(has_problem, diff_text)
    """
    alloc = alloc or {}
    used = used or {}

    codes = sorted(set(list(alloc.keys()) + list(used.keys())))
    diffs = []

    for c in codes:
        a = int(alloc.get(c, 0))
        u = int(used.get(c, 0))
        if a < u:
            diffs.append(f"{c}:151={a},QD={u},不足={u-a}")

    if diffs:
        return True, "；".join(diffs)
    return False, ""

def calc_actual_ba_by_case_and_month(sess: requests.Session, case_id: str, roc_yyyymm: int, org_filter: str | None = None) -> tuple[dict, str]:
    """
    PageE 用：計算單一 case_id 在單月內，QD120+QD120A 的 BA 實際使用量
    - 新增 org_filter：只計入「長照機構=選定機構」的 row
    回傳：(ba_map, note)
    """
    servDt1_enc, servDt2_enc, _human = roc_yyyymm_to_servDt_enc(roc_yyyymm)

    rows1, n1 = qd_fetch_all_pages(sess, QD120_URL_TEMPLATE, case_id, servDt1_enc, servDt2_enc, limit=100)
    for r in rows1:
        r["_source"] = "filterQd120"
    rows2, n2 = qd_fetch_all_pages(sess, QD120A_URL_TEMPLATE, case_id, servDt1_enc, servDt2_enc, limit=100)
    for r in rows2:
        r["_source"] = "filterQd120A"

    merged = dedup_rows(rows1 + rows2)

    # ✅ 機構篩選（不過濾則全留）
    if org_filter and _norm_text(org_filter) != _norm_text(ORG_NO_FILTER):
        merged = [r for r in merged if row_matches_org(r, org_filter)]

    ba_map: dict[str, int] = {}
    for r in merged:
        code = extract_code_any(r)
        if not code.startswith("BA"):
            continue
        ba_map[code] = ba_map.get(code, 0) + int(get_serv_count_guess(r))

    note_parts = []
    if n1:
        note_parts.append(f"QD120:{n1}")
    if n2:
        note_parts.append(f"QD120A:{n2}")
    if org_filter and _norm_text(org_filter) != _norm_text(ORG_NO_FILTER):
        note_parts.append(f"機構篩選:{org_filter}")

    note = "；".join(note_parts) if note_parts else ""
    return ba_map, note


# =========================================================
# ✅【補齊】PageD 缺少的函式：用 QD120/QD120A 算區間內 GA09 實際使用量
# =========================================================
def calc_actual_ga09_by_name_and_range(sess: requests.Session, person: str, roc_start_yyyymm: int, roc_end_yyyymm: int, org_filter: str | None = None) -> tuple[int, list[str], str]:
    """
    PageD 用：
    - 先用姓名查 case_id（嚴格同名；多案會加總）
    - 再抓 QD120 + QD120A 在區間內的紀錄
    - 只統計 GA09（視為喘息 GA09）
    回傳：(ga09_sum, case_ids, note)
    """
    case_ids, note_ca = api_case_find_case_ids_strict(sess, person)
    if not case_ids:
        return 0, [], note_ca or "查無 case_id"

    servDt1_enc, servDt2_enc, _human = roc_yyyymm_range_to_servDt_enc(roc_start_yyyymm, roc_end_yyyymm)

    total_ga09 = 0
    qd_notes = []

    for cid in case_ids:
        rows1, n1 = qd_fetch_all_pages(sess, QD120_URL_TEMPLATE, cid, servDt1_enc, servDt2_enc, limit=100)
        for r in rows1:
            r["_source"] = "filterQd120"
        rows2, n2 = qd_fetch_all_pages(sess, QD120A_URL_TEMPLATE, cid, servDt1_enc, servDt2_enc, limit=100)
        for r in rows2:
            r["_source"] = "filterQd120A"

        merged = dedup_rows(rows1 + rows2)

        # ✅ 機構篩選（不過濾則全留）
        if org_filter and _norm_text(org_filter) != _norm_text(ORG_NO_FILTER):
            merged = [r for r in merged if row_matches_org(r, org_filter)]

        ga09 = 0
        for r in merged:
            code = extract_code_any(r)
            # 保守：GA09 / GA09-? 都算
            if code == "GA09" or code.startswith("GA09"):
                ga09 += int(get_serv_count_guess(r))

        total_ga09 += ga09

        tmp = []
        if n1:
            tmp.append(f"QD120:{n1}")
        if n2:
            tmp.append(f"QD120A:{n2}")
        if org_filter and _norm_text(org_filter) != _norm_text(ORG_NO_FILTER):
            tmp.append(f"機構篩選:{org_filter}")
        if tmp:
            qd_notes.append(f"{cid}({'/'.join(tmp)})")

    note_parts = []
    if note_ca:
        note_parts.append(note_ca)
    if qd_notes:
        note_parts.append("；".join(qd_notes))

    return total_ga09, case_ids, "；".join(note_parts).strip("；")


# =========================================================
# PageA：GA09 名單：從「服務紀錄總表單」抓 GA09 / 個案 / 排班日期
# =========================================================
def build_ga09_list_from_service_total(excel_path: str, month_str: str) -> pd.DataFrame:
    """
    產出欄位：
      月份、個案姓名、GA09次數、最後一次排班日期(YYYY-MM-DD)
        條件：
            只計入「服務項目代碼=GA09」且「狀態(U欄)=已簽退」的資料列
    """
    m = str(month_str).strip().zfill(2)
    try:
        m_int = int(m)
        if not (1 <= m_int <= 12):
            raise ValueError()
    except Exception:
        raise ValueError("輸出月份需為 01~12")

    raw = pd.read_excel(excel_path, header=None)

    # C=2, F=5, U=20, V=21；資料從第5列開始（index=4）
    start_row = 4
    code_col = 2
    name_col = 5
    status_col = 20
    date_col = 21

    df = raw.iloc[start_row:, [code_col, name_col, status_col, date_col]].copy()
    df.columns = ["服務項目代碼", "個案姓名", "狀態", "排班日期"]

    df["服務項目代碼"] = df["服務項目代碼"].astype(str).str.strip().str.upper()
    df["個案姓名"] = df["個案姓名"].astype(str).str.strip()
    df["狀態"] = df["狀態"].astype(str).str.replace(r"[\s　]+", "", regex=True)

    dt = pd.to_datetime(df["排班日期"], errors="coerce")
    if dt.isna().all():
        num = pd.to_numeric(df["排班日期"], errors="coerce")
        dt = pd.to_datetime(num, unit="D", origin="1899-12-30", errors="coerce")
    df["排班日期_dt"] = dt

    df = df[
        (df["服務項目代碼"] == "GA09")
        & (df["狀態"] == "已簽退")
        & (df["排班日期_dt"].notna())
    ]
    if df.empty:
        return pd.DataFrame(columns=["月份", "個案姓名", "GA09次數", "最後一次排班日期"])

    df_m = df[df["排班日期_dt"].dt.month == int(m)].copy()
    if df_m.empty:
        return pd.DataFrame(columns=["月份", "個案姓名", "GA09次數", "最後一次排班日期"])

    years = df_m["排班日期_dt"].dt.year
    year_mode = int(years.mode().iloc[0]) if not years.mode().empty else int(years.max())
    df_m = df_m[df_m["排班日期_dt"].dt.year == year_mode].copy()

    g = df_m.groupby("個案姓名", dropna=False).agg(
        GA09次數=("服務項目代碼", "size"),
        最後一次排班日期=("排班日期_dt", "max")
    ).reset_index()

    g["最後一次排班日期"] = g["最後一次排班日期"].dt.strftime("%Y-%m-%d")
    g.insert(0, "月份", f"{m}月")
    g["GA09次數"] = g["GA09次數"].astype(int)

    g = g.sort_values(by=["GA09次數", "最後一次排班日期"], ascending=[False, False]).reset_index(drop=True)
    return g


# =========================================================
# 151 API（QP111）
# =========================================================
def api_qp111_fetch(sess: requests.Session, name: str, yyy: str, mm: str, offset: int = 0, limit: int = 200) -> dict:
    params = {
        "qdperms": "true",
        "caseno": "",
        "serialno": "",
        "name": name,
        "idno": "",
        "qcntcode": "",
        "twnspcode": "",
        "vilgcode": "",
        "qinformCntcode": "",
        "informTwnspcode": "",
        "informVilgcode": "",
        "yyy": yyy,
        "mm": mm,
        "_adjHint": "",
        "_hasPi400a": "",
        "_aliveFalse": "",
        "order": "asc",
        "offset": offset,
        "limit": limit,
    }
    return http_get_json_with_retry(
        sess,
        API_QP111_FILTER,
        params=params,
        headers=REQ_HEADERS,
        timeout=30,
        max_retries=3,
    )

def api_qp111_query_one_person_one_month(sess: requests.Session, person: str, ym: str) -> list[dict]:
    yyy = ym[:-2]
    mm = ym[-2:]
    mm_int = str(int(mm))

    out = []
    js = api_qp111_fetch(sess, name=person, yyy=yyy, mm=mm, offset=0, limit=200)
    rows = js.get("rows", []) or []
    out.extend([r for r in rows if str(r.get("yyymm", "")).strip() == ym and str(r.get("name", "")).strip() == person])

    if not out and mm_int != mm:
        js2 = api_qp111_fetch(sess, name=person, yyy=yyy, mm=mm_int, offset=0, limit=200)
        rows2 = js2.get("rows", []) or []
        out.extend([r for r in rows2 if str(r.get("yyymm", "")).strip() == ym and str(r.get("name", "")).strip() == person])

    return out


# =========================================================
# 152 API（QP211：區間）
# =========================================================
def api_qp211_fetch(sess: requests.Session, name: str, yyy: str = "", offset: int = 0, limit: int = 200) -> dict:
    params = {
        "qdperms": "true",
        "caseno": "",
        "serialno": "",
        "name": name,
        "idno": "",
        "qcntcode": "",
        "twnspcode": "",
        "vilgcode": "",
        "qinformCntcode": "",
        "informTwnspcode": "",
        "informVilgcode": "",
        "yyy": yyy,
        "_adjHint": "",
        "_hasPi400a": "",
        "_aliveFalse": "",
        "order": "asc",
        "offset": offset,
        "limit": limit,
    }
    return http_get_json_with_retry(
        sess,
        API_QP211_FILTER,
        params=params,
        headers=REQ_HEADERS,
        timeout=30,
        max_retries=3,
    )

def api_qp211_query_person_matched_rows(sess: requests.Session, person: str, target_ym: str) -> list[dict]:
    """
    回傳：QP211 rows（只保留 name=person 且 yyymm 區間包含 target_ym）
    """
    js = api_qp211_fetch(sess, name=person, yyy="", offset=0, limit=200)
    rows = js.get("rows", []) or []
    matched = []
    for r in rows:
        if str(r.get("name", "")).strip() != person:
            continue
        rng = str(r.get("yyymm", "")).strip()
        if ym_in_range(target_ym, rng):
            matched.append(r)
    return matched

def summarize_qp211_matched(matched_rows: list[dict]) -> tuple[int | None, list[str], str, str]:
    """
    將 matched rows 彙整：
    回傳：
      (ga09_sum_or_none, ranges_list, latest_lastUpdated_text, note)
    """
    if not matched_rows:
        return (None, [], "", "查無資料或區間不包含目標月")

    ga09_sum = 0
    ranges = []
    latest_ym = None
    latest_lu = ""

    for r in matched_rows:
        rng = str(r.get("yyymm", "")).strip()
        if rng and (rng not in ranges):
            ranges.append(rng)

        qd = (r.get("qd100s", "") or "")
        m = parse_qd_pairs(qd)
        ga09_sum += int(m.get("GA09", 0))

        lu = str(r.get("lastUpdated", "")).strip()
        lu_ym = roc_date_to_ym(lu)
        if lu_ym and lu_ym.isdigit():
            if (latest_ym is None) or (int(lu_ym) > int(latest_ym)):
                latest_ym = lu_ym
                latest_lu = lu
        elif not latest_lu and lu:
            latest_lu = lu

    return (ga09_sum, ranges, latest_lu, "")


# =========================================================
# UI：首頁圓角卡片
# =========================================================
def draw_round_rect(canvas: tk.Canvas, x1, y1, x2, y2, r=18, **kw):
    points = [
        x1+r, y1, x2-r, y1, x2, y1, x2, y1+r,
        x2, y2-r, x2, y2, x2-r, y2, x1+r, y2,
        x1, y2, x1, y2-r, x1, y1+r, x1, y1
    ]
    return canvas.create_polygon(points, smooth=True, **kw)


class HomePage(tk.Frame):
    def __init__(self, master, goto_a, goto_b, goto_c, goto_d, goto_e, goto_adv, clear_session_cb, shared_user_var, shared_pwd_var):
        super().__init__(master, bg=BG_MAIN)
        self.goto_a, self.goto_b, self.goto_c, self.goto_d, self.goto_e = goto_a, goto_b, goto_c, goto_d, goto_e
        self.goto_adv = goto_adv
        self.clear_session_cb = clear_session_cb
        self.shared_user_var = shared_user_var
        self.shared_pwd_var = shared_pwd_var
        self.pack(fill="both", expand=True)
        self.build()

    def build(self):
        W, H = 900, 720
        card_w, card_h = 660, 80
        gap_y = 28

        canvas = tk.Canvas(self, width=W, height=H, bg=BG_MAIN, highlightthickness=0)
        canvas.pack(fill="both", expand=True)

        canvas.create_text(W//2, 45, text="額分次數查詢系統",
                           fill=FG_MAIN, font=("Microsoft JhengHei UI", 26, "bold"))

        btn = tk.Button(
            self, text="清除登入快取/換帳號",
            command=self.clear_session_cb,
            bg="#B91C1C", fg="white", relief="flat", cursor="hand2"
        )
        btn.place(x=W-200, y=20, width=170, height=32)

        tk.Label(self, text="LCMS 帳號：", bg=BG_MAIN, fg=FG_MAIN).place(x=120, y=80)
        tk.Entry(self, textvariable=self.shared_user_var, width=22,
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=220, y=80)

        tk.Label(self, text="LCMS 密碼：", bg=BG_MAIN, fg=FG_MAIN).place(x=440, y=80)
        tk.Entry(self, textvariable=self.shared_pwd_var, width=22, show="*",
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=540, y=80)

        def make_card(y, text, cb):
            x = (W - card_w)//2
            box = draw_round_rect(canvas, x, y, x+card_w, y+card_h, r=22,
                                  fill=BG_CARD, outline="#444444", width=2)
            txt = canvas.create_text(x+30, y+card_h/2, anchor="w",
                                     text=text, fill=FG_MAIN,
                                     font=("Microsoft JhengHei UI", 17, "bold"))

            def enter(_): canvas.itemconfig(box, fill="#2A2A2A")
            def leave(_): canvas.itemconfig(box, fill=BG_CARD)
            def click(_): cb()

            for item in (box, txt):
                canvas.tag_bind(item, "<Enter>", enter)
                canvas.tag_bind(item, "<Leave>", leave)
                canvas.tag_bind(item, "<Button-1>", click)

        y0 = 120
        make_card(y0,                          "① 額分校對（一鍵流程）", self.goto_a)
        make_card(y0 + (card_h + gap_y)*1,     "② 151名單 VS 當月服務紀錄比對（BA）", self.goto_e)
        make_card(y0 + (card_h + gap_y)*2,     "進階功能", self.goto_adv)

        canvas.create_text(W-20, H-23, anchor="se",
                           text="FOR M. BY BEN",
                           fill="#666666", font=("Microsoft JhengHei UI", 10))


# =========================================================
# PageA（主名單 + GA09 名單，輸出兩份 Excel）
# =========================================================
class PageA(tk.Frame):
    def __init__(self, master, go_home):
        super().__init__(master, bg=BG_CARD)
        self.go_home = go_home
        self.file_path = tk.StringVar()
        self.month_str = tk.StringVar()
        self.target_ym = tk.StringVar()
        self.build()

        today = datetime.today()
        last_month = today.replace(day=1) - timedelta(days=1)
        self.month_str.set(str(last_month.month).zfill(2))

        roc_year = today.year - 1911
        self.target_ym.set(f"{roc_year}{self.month_str.get()}")

    def build(self):
        tk.Button(self, text="← 回首頁", command=self.go_home,
                  bg=BG_CARD, fg=FG_MAIN, relief="flat", cursor="hand2",
                  activebackground=BG_CARD, activeforeground=FG_MAIN).place(x=20, y=15)

        tk.Label(self, text="仁寶當月服務數量查詢（輸出主名單 + GA09 喘息名單）",
                 font=("Microsoft JhengHei UI", 20, "bold"),
                 bg=BG_CARD, fg=FG_MAIN).place(relx=0.5, y=35, anchor="center")

        tk.Button(self, text="選擇仁寶服務紀錄 Excel", command=self.choose_file,
                  bg="#4B5563", fg="white",
                  activebackground="#6B7280", activeforeground="white",
                  cursor="hand2").place(x=60, y=120)

        tk.Label(self, textvariable=self.file_path, fg=FG_LINK, bg=BG_CARD).place(x=270, y=120)

        tk.Label(self, text="輸出月份：", bg=BG_CARD, fg=FG_MAIN).place(x=60, y=150)
        tk.Entry(self, textvariable=self.month_str, width=6,
                 bg=BG_TEXT, fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=180, y=150)

        tk.Label(self, text="151分配年月：", bg=BG_CARD, fg=FG_MAIN).place(x=280, y=150)
        tk.Entry(self, textvariable=self.target_ym, width=10,
             bg=BG_TEXT, fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=390, y=150)
        tk.Label(self, text="例：11503", bg=BG_CARD, fg=FG_HINT).place(x=500, y=150)

        tk.Button(self, text="一鍵執行並輸出結果", width=30, height=2,
                  command=lambda: threading.Thread(target=self.run, daemon=True).start(),
                  bg=ACCENT_GO, fg=FG_MAIN, activebackground="#303F9F",
                  activeforeground="#6B7280", cursor="hand2").place(x=60, y=180)

        self.progress = tk.IntVar(value=0)
        self.bar = ttk.Progressbar(self, variable=self.progress, length=820)
        self.bar.place(x=60, y=220)

        self.log = scrolledtext.ScrolledText(self, width=96, height=20,
                                             bg=BG_TEXT, fg=FG_MAIN,
                                             insertbackground=FG_MAIN, relief="flat")
        self.log.place(x=30, y=250, width=840, height=420)

        tk.Label(self, text="FOR M. BY BEN", bg=self["bg"], fg="#666666",
                 font=("Microsoft JhengHei UI", 10)).place(relx=1.0, rely=1.0, x=-20, y=-20, anchor="se")

    def choose_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.file_path.set(path)

    def _log(self, msg: str):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)

    def _set_progress(self, pct: int):
        val = max(0, min(100, int(pct)))
        self.progress.set(val)
        self.bar.update()

    def _stage_progress(self, start: int, end: int, inner_pct: int):
        v = max(0, min(100, int(inner_pct)))
        self._set_progress(int(start + (end - start) * v / 100))

    def run(self):
        self.log.delete(1.0, tk.END)
        self._set_progress(0)
        if not self.file_path.get():
            messagebox.showerror("錯誤", "請先選擇 Excel")
            return

        m = self.month_str.get().zfill(2)
        target_ym = self.target_ym.get().strip()
        if len(target_ym) != 5 or (not target_ym.isdigit()):
            messagebox.showerror("錯誤", "請輸入正確的 151分配年月，例如：11503")
            return

        self._log("🚀 開始一鍵流程（①+②+④+③）")
        self._stage_progress(0, 20, 5)

        # --------
        # 主名單（沿用你原本邏輯）
        # --------
        df_out = None
        try:
            df = pd.read_excel(self.file_path.get(), header=2)

            unit_col = df.columns[49]      # AX
            manager_col = df.columns[50]   # AY

            status = df.columns[20]
            df[status] = df[status].astype(str).str.replace(r"[\s　]+", "", regex=True)
            df = df[df[status].astype(str).str.contains("已簽退|已簽到|週期排班|單次排班", na=False)]

            pay = df.columns[7]
            df[pay] = df[pay].astype(str).str.replace(r"\s+", "", regex=True)
            df["本筆費用別"] = df[pay].apply(lambda x: "自費" if "自費" in x else "補助")

            fee_map = df.groupby("姓名")["本筆費用別"].apply(
                lambda x: "補助+自費" if set(x) == {"補助", "自費"} else x.iloc[0]
            )
            df["費用來源"] = df["姓名"].map(fee_map)

            df["項目服務次數"] = pd.to_numeric(df["項目服務次數"], errors="coerce").fillna(0)
            summary = df.groupby(["姓名", "服務項目代碼", "本筆費用別"])["項目服務次數"].sum().reset_index()

            unit_map = df.groupby("姓名")[unit_col].first()
            manager_map = df.groupby("姓名")[manager_col].first()

            output = []
            for name, group in summary.groupby("姓名"):
                items = []
                for code, sub in group.groupby("服務項目代碼"):
                    s = sub[sub["本筆費用別"] == "補助"]["項目服務次數"].sum()
                    p = sub[sub["本筆費用別"] == "自費"]["項目服務次數"].sum()
                    if s > 0 and p > 0:
                        items.append(f"{code}：{int(s)}(補助)/{int(p)}(自費)")
                    elif s > 0:
                        items.append(f"{code}：{int(s)}")
                    elif p > 0:
                        items.append(f"{code}：{int(p)}(自費)")

                output.append([
                    f"{m}月", name, fee_map[name], "、".join(items),
                    unit_map.get(name, ""), manager_map.get(name, "")
                ])

            df_out = pd.DataFrame(output, columns=["月份", "姓名", "費用來源", "服務內容", "A單位名稱", "A個管姓名"])
            self._log("✅ 主名單彙整完成")
        except Exception as e:
            self._log(f"⚠ 主名單彙整失敗（你原本格式可能不同）：{e}")
            df_out = pd.DataFrame(columns=["月份", "姓名", "費用來源", "服務內容", "A單位名稱", "A個管姓名"])

        # --------
        # GA09 喘息名單
        # --------
        try:
            df_ga09 = build_ga09_list_from_service_total(self.file_path.get(), m)
            self._log(f"✅ GA09 喘息名單彙整完成：{len(df_ga09)} 位")
        except Exception as e:
            self._log(f"⚠ GA09 喘息名單彙整失敗：{e}")
            df_ga09 = pd.DataFrame(columns=["月份", "個案姓名", "GA09次數", "最後一次排班日期"])

        self._stage_progress(0, 20, 55)

        # -----------------------------
        # 輸出：兩份 Excel（不是分頁）
        # -----------------------------
        save_main = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"{m}月名單.xlsx",
            filetypes=[("Excel 檔案", "*.xlsx")]
        )
        if not save_main:
            messagebox.showwarning("未存檔", "你取消了存檔")
            return

        import os
        base_dir = os.path.dirname(save_main)
        base_name = os.path.splitext(os.path.basename(save_main))[0]
        default_ga09 = os.path.join(base_dir, f"{base_name}_GA09喘息名單.xlsx")

        save_ga09 = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=os.path.basename(default_ga09),
            initialdir=base_dir,
            filetypes=[("Excel 檔案", "*.xlsx")]
        )
        if not save_ga09:
            messagebox.showwarning("未存檔", "你取消了 GA09 名單存檔（主名單仍會輸出）")
            save_ga09 = None

        try:
            df_out.to_excel(save_main, index=False)
            if save_ga09:
                df_ga09.to_excel(save_ga09, index=False)
        except Exception as e:
            messagebox.showerror("錯誤", f"輸出失敗：\n{e}")
            return

        self.master.last_local_path = save_main
        if save_ga09:
            self.master.last_ga09_path = save_ga09
        self.master.last_target_ym = target_ym
        self._stage_progress(0, 20, 100)

        self._log(f"🎉 ①完成 主名單：{save_main}")
        if save_ga09:
            self._log(f"🎉 ①完成 GA09 喘息名單：{save_ga09}")

        # -------------------------------------------------
        # 一鍵串接：② 151 額分查詢 + ④ 152/喘息校對 + ③ 需補分配比對
        # -------------------------------------------------
        qp111_auto = os.path.join(base_dir, f"{base_name}_151額分.xlsx")
        qp211_auto = os.path.join(base_dir, f"{base_name}_152_區間vs實際(GA09).xlsx")
        cmp_auto = os.path.join(base_dir, f"{base_name}_151需補分配.xlsx")

        self._log("🚀 開始自動執行 ② 151 額分查詢...")

        self.master.page_b.selected_file.set(save_main)
        self.master.page_b.month.set(target_ym)
        try:
            if hasattr(self.master, "shared_username") and self.master.shared_username.get().strip():
                self.master.page_b.username.set(self.master.shared_username.get().strip())
            if hasattr(self.master, "shared_password") and self.master.shared_password.get().strip():
                self.master.page_b.password.set(self.master.shared_password.get().strip())
        except Exception:
            pass

        ok_b = self.master.page_b.run_query(
            save_path=qp111_auto,
            auto_mode=True,
            log_hook=lambda m: self._log(f"[②] {m}"),
            progress_hook=lambda p: self._stage_progress(20, 50, p)
        )
        if not ok_b:
            self._log("❌ ② 執行失敗，已停止後續流程。")
            return

        self._log(f"✅ ② 完成：{qp111_auto}")

        selected_org = ui_blocking_select_option(
            self.master,
            "選擇喘息單位",
            "請選擇要用於④『152+實際喘息校對』的長照機構：",
            ORG_OPTIONS,
            default=self.master.page_d.org_var.get().strip() if hasattr(self.master, "page_d") else ORG_OPTIONS[0]
        )
        if selected_org is None:
            self._log("⚠ 你取消了喘息單位選擇，已停止後續流程。")
            return

        self._log(f"🚀 開始自動執行 ④ 152+實際喘息校對（機構：{selected_org}）...")

        self.master.page_d.selected_file.set(save_ga09 if save_ga09 else default_ga09)
        self.master.page_d.target_ym.set(target_ym)
        self.master.page_d.org_var.set(selected_org)
        try:
            if hasattr(self.master, "shared_username") and self.master.shared_username.get().strip():
                self.master.page_d.username.set(self.master.shared_username.get().strip())
            if hasattr(self.master, "shared_password") and self.master.shared_password.get().strip():
                self.master.page_d.password.set(self.master.shared_password.get().strip())
        except Exception:
            pass

        ga09_source = save_ga09 if (save_ga09 and os.path.exists(save_ga09)) else ""
        if not ga09_source:
            self._log("❌ 找不到 GA09 名單檔案，請不要取消 GA09 存檔。")
            return
        self.master.page_d.selected_file.set(ga09_source)

        need_step4 = True
        try:
            df_ga09_for_step4 = pd.read_excel(ga09_source, dtype=str).fillna("")
            if "個案姓名" not in df_ga09_for_step4.columns:
                raw = pd.read_excel(ga09_source, header=None)
                tmp = raw.iloc[1:, [1]].copy()
                tmp.columns = ["個案姓名"]
                df_ga09_for_step4 = tmp.fillna("")

            has_name = df_ga09_for_step4["個案姓名"].astype(str).str.strip().ne("").any()
            need_step4 = bool(has_name)
        except Exception:
            need_step4 = True

        ran_step4 = False
        respite_rows = []
        if need_step4:
            ok_d = self.master.page_d.run_query(
                save_path=qp211_auto,
                auto_mode=True,
                org_selected_override=selected_org,
                log_hook=lambda m: self._log(f"[④] {m}"),
                progress_hook=lambda p: self._stage_progress(50, 80, p)
            )
            if not ok_d:
                self._log("❌ ④ 執行失敗，已停止後續流程。")
                return

            respite_rows = list(getattr(self.master.page_d, "results", []) or [])
            ran_step4 = True
            self._log(f"✅ ④ 完成：{qp211_auto}")
        else:
            self._stage_progress(50, 80, 100)
            self._log("ℹ GA09 名單沒有姓名，已略過 ④ 152+實際喘息校對。")
        self._log("🚀 開始自動執行 ③ 需補分配比對（含喘息補分配）...")

        self.master.page_c.file_local.set(save_main)
        self.master.page_c.file_qp111.set(qp111_auto)

        ok_c = self.master.page_c.run(
            save_path=cmp_auto,
            auto_mode=True,
            respite_rows=respite_rows,
            log_hook=lambda m: self._log(f"[③] {m}"),
            progress_hook=lambda p: self._stage_progress(80, 100, p)
        )
        if not ok_c:
            self._log("❌ ③ 執行失敗，請查看 PageC 記錄。")
            return

        cmp_saved = str(getattr(self.master.page_c, "last_output_path", "") or "").strip()
        no_need_alloc = bool(getattr(self.master.page_c, "last_no_need_allocation", False))
        if cmp_saved:
            self._log(f"✅ ③ 完成：{cmp_saved}")
        elif no_need_alloc:
            self._log("✅ ③ 完成：無需分配（未輸出檔案）")
        else:
            self._log("✅ ③ 完成")
        self._set_progress(100)
        self._log("🎉 一鍵流程完成（1+2+4+3）")
        self._log(f"1) 主名單：{save_main}")
        self._log(f"2) 151額分：{qp111_auto}")
        if ran_step4:
            self._log(f"4) 152喘息校對：{qp211_auto}")
        else:
            self._log("4) 152喘息校對：略過（GA09 名單無姓名）")
        if cmp_saved:
            self._log(f"3) 需補分配(含喘息)：{cmp_saved}")
        elif no_need_alloc:
            self._log("3) 需補分配(含喘息)：無需分配")
        else:
            self._log("3) 需補分配(含喘息)：未輸出")
        if ran_step4:
            messagebox.showinfo("完成", "已完成一鍵流程（1+2+4+3）")
        else:
            messagebox.showinfo("完成", "已完成一鍵流程（1+2+③，④因 GA09 名單無姓名而略過）")


# =========================================================
# PageB：151（QP111）
# =========================================================
class PageB(tk.Frame):
    def __init__(self, master, go_home):
        super().__init__(master, bg=BG_CARD)
        self.go_home = go_home

        self.username = tk.StringVar()
        self.password = tk.StringVar()
        self.selected_file = tk.StringVar()
        self.month = tk.StringVar()
        self.progress = tk.IntVar()

        self.session = None
        self.results = []

        self.build()

    def _ui_log(self, msg: str):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)

    def append_log(self, msg: str):
        self.after(0, lambda: self._ui_log(msg))

    def update_progress(self, now, total):
        def _upd():
            self.progress.set(0 if total <= 0 else int(now / total * 100))
            self.bar.update()
        self.after(0, _upd)

    def tree_clear(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

    def tree_add_row(self, row: dict):
        self.tree.insert(
            "", "end",
            values=(
                row.get("姓名", ""),
                row.get("分配年月", ""),
                row.get("額分(BA碼:次數)", ""),
                row.get("備註", "")
            )
        )

    def _get_or_login_session(self) -> requests.Session | None:
        if getattr(self.master, "shared_session", None):
            self.session = self.master.shared_session
            return self.session
        if self.session:
            return self.session

        user = self.username.get().strip() or getattr(self.master, "shared_username", tk.StringVar()).get().strip()
        pwd = self.password.get().strip() or getattr(self.master, "shared_password", tk.StringVar()).get().strip()
        if not user or not pwd:
            messagebox.showerror("錯誤", "請先輸入 LCMS 帳號與密碼")
            return None

        self.username.set(user)
        self.password.set(pwd)
        if hasattr(self.master, "shared_username"):
            self.master.shared_username.set(user)
        if hasattr(self.master, "shared_password"):
            self.master.shared_password.set(pwd)

        self.append_log("🔐 正在登入 LCMS（驗證碼/Cloudflare/reCAPTCHA 需人工處理，皆在程式內提示）...")
        sess = lcms_login_gui(self.master, user, pwd)
        if not sess:
            return None
        self.session = sess
        self.master.shared_session = sess
        return sess

    def run_query(self, save_path: str | None = None, auto_mode: bool = False,
                  log_hook=None, progress_hook=None):
        self.btn_start.config(state="disabled")
        self.log.delete(1.0, tk.END)
        self.tree_clear()
        self.results = []

        def emit(msg: str):
            self.append_log(msg)
            if callable(log_hook):
                try:
                    log_hook(msg)
                except Exception:
                    pass

        ym = self.month.get().strip()
        if len(ym) != 5 or (not ym.isdigit()):
            messagebox.showerror("錯誤", "請輸入正確的『分配年月』，例如：11412")
            self.btn_start.config(state="normal")
            return False

        file_path = self.selected_file.get()
        if not file_path:
            messagebox.showerror("錯誤", "請先選擇名單檔案")
            self.btn_start.config(state="normal")
            return False

        try:
            df = pd.read_excel(file_path, header=None)
            names = df.iloc[1:, 1].dropna().astype(str).str.strip().tolist()
        except Exception as e:
            messagebox.showerror("錯誤", f"名單讀取失敗（請確認 Excel：B欄為姓名）\n{e}")
            self.btn_start.config(state="normal")
            return False

        if not names:
            messagebox.showwarning("提示", "名單內沒有任何姓名")
            self.btn_start.config(state="normal")
            return False

        self.lbl_current_month.config(text=f"目前查詢：{ym}（151 額分）")
        emit(f"🟢 名單共 {len(names)} 位")
        emit(f"🟢 分配年月：{ym}")
        emit("🟢 顯示內容：該月 151 額分（BA碼:次數）\n")

        sess = self._get_or_login_session()
        if not sess:
            emit("❌ 登入失敗或取消。")
            self.btn_start.config(state="normal")
            return False
        try:
            setattr(sess, "_retry_log_hook", emit)
        except Exception:
            pass

        emit("✅ 登入成功/已共用連線，開始查詢...\n")

        total = len(names)

        for idx, person in enumerate(names, start=1):
            emit(f"[{idx}/{total}] {person}")

            try:
                matched_rows = api_qp111_query_one_person_one_month(sess, person, ym)

                if not matched_rows:
                    row_out = {
                        "分配年月": ym,
                        "姓名": person,
                        "額分(BA碼:次數)": "",
                        "備註": "查無資料"
                    }
                    self.results.append(row_out)
                    self.after(0, lambda r=row_out: self.tree_add_row(r))
                    emit("   ❌ 查無該月分配資料")
                else:
                    merged = {}
                    for r in matched_rows:
                        qd100s = (r.get("qd100s", "") or "").strip()
                        alloc_map = parse_qd_pairs(qd100s)
                        for k, v in alloc_map.items():
                            merged[k] = merged.get(k, 0) + int(v)

                    merged_ba = {k: v for k, v in merged.items() if str(k).upper().startswith("BA")}
                    alloc_text = "、".join([f"{k}:{v}" for k, v in sorted(merged_ba.items())]) if merged_ba else ""

                    row_out = {
                        "分配年月": ym,
                        "姓名": person,
                        "額分(BA碼:次數)": alloc_text,
                        "備註": "" if alloc_text else "有資料但無可辨識BA碼"
                    }
                    self.results.append(row_out)
                    self.after(0, lambda rr=row_out: self.tree_add_row(rr))
                    emit("   ✅ 已抓到（同名/多案會自動加總）")
                    if alloc_text:
                        emit(f"      {alloc_text}")

            except Exception as e:
                row_out = {
                    "分配年月": ym,
                    "姓名": person,
                    "額分(BA碼:次數)": "",
                    "備註": f"異常：{e}"
                }
                self.results.append(row_out)
                self.after(0, lambda r=row_out: self.tree_add_row(r))
                emit(f"   ❌ 查詢異常：{e}")

            self.update_progress(idx, total)
            if callable(progress_hook):
                try:
                    progress_hook(int(idx * 100 / total))
                except Exception:
                    pass

        if not save_path:
            save_path = filedialog.asksaveasfilename(
                title="另存 151 額分查詢結果",
                defaultextension=".xlsx",
                initialfile=f"{ym}_151額分.xlsx",
                filetypes=[("Excel 檔案", "*.xlsx")]
            )
        if save_path:
            try:
                pd.DataFrame(self.results).to_excel(save_path, index=False)
                self.master.last_qp111_path = save_path
                self.master.last_list_path = file_path
                self.master.last_target_ym = ym
                if not auto_mode:
                    messagebox.showinfo("完成", f"已儲存：\n{save_path}")
                emit(f"\n🎉 完成！已輸出：{save_path}")
            except Exception as e:
                messagebox.showerror("錯誤", f"儲存 Excel 失敗：\n{e}")
                self.btn_start.config(state="normal")
                return False
        else:
            if not auto_mode:
                messagebox.showwarning("未儲存", "你取消了存檔（畫面仍保留查詢結果）")
            self.btn_start.config(state="normal")
            return False

        self.btn_start.config(state="normal")
        if callable(progress_hook):
            try:
                progress_hook(100)
            except Exception:
                pass
        return True

    def build(self):
        tk.Button(
            self, text="← 回首頁", command=self.go_home,
            bg=BG_CARD, fg=FG_MAIN, relief="flat", cursor="hand2"
        ).place(x=20, y=15)

        tk.Label(
            self, text="151 額分查詢",
            font=("Microsoft JhengHei UI", 22, "bold"),
            bg=BG_CARD, fg=FG_MAIN
        ).place(relx=0.5, y=20, anchor="center")

        self.lbl_current_month = tk.Label(
            self, text="目前查詢：—",
            font=("Microsoft JhengHei UI", 11),
            bg=BG_CARD, fg=FG_HINT
        )
        self.lbl_current_month.place(relx=0.5, y=55, anchor="center")

        tk.Label(self, text="LCMS 帳號：", bg=BG_CARD, fg=FG_MAIN).place(x=60, y=80)
        tk.Entry(self, textvariable=self.username, width=18,
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=150, y=80)

        tk.Label(self, text="LCMS 密碼：", bg=BG_CARD, fg=FG_MAIN).place(x=340, y=80)
        tk.Entry(self, textvariable=self.password, width=18, show="*",
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=430, y=80)

        tk.Button(
            self, text="選擇月份名單",
            command=self.choose, width=18,
            bg="#4B5563", fg="white", cursor="hand2"
        ).place(x=60, y=115)
        tk.Label(self, textvariable=self.selected_file, fg=FG_LINK, bg=BG_CARD).place(x=245, y=118)

        tk.Label(self, text="分配年月：", bg=BG_CARD, fg=FG_MAIN).place(x=60, y=155)
        tk.Entry(self, textvariable=self.month, width=10,
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=150, y=155)
        tk.Label(self, text="例：11412", bg=BG_CARD, fg=FG_HINT).place(x=245, y=155)

        self.btn_start = tk.Button(
            self, text="開始查詢", width=12, height=1,
            command=lambda: threading.Thread(target=self.run_query, daemon=True).start(),
            bg=ACCENT_OK, fg="white",
            font=("Microsoft JhengHei UI", 12, "bold"),
            cursor="hand2"
        )
        self.btn_start.place(x=750, y=150)

        self.bar = ttk.Progressbar(self, variable=self.progress, length=820)
        self.bar.place(x=60, y=190)

        cols = ("姓名", "分配年月", "額分(BA碼:次數)", "備註")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=9)
        for c in cols:
            self.tree.heading(c, text=c)

        self.tree.tag_configure("insufficient", foreground="#FF4D4F")

        self.tree.column("姓名", width=110, anchor="w")
        self.tree.column("分配年月", width=80, anchor="center")
        self.tree.column("額分(BA碼:次數)", width=520, anchor="w")
        self.tree.column("備註", width=90, anchor="w")
        self.tree.place(x=60, y=215, width=820, height=250)

        self.log = scrolledtext.ScrolledText(
            self, width=117, height=10,
            bg="#111111", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat"
        )
        self.log.place(x=60, y=480)

        tk.Label(self, text="FOR M. BY BEN", bg=BG_CARD, fg="#777777",
                 font=("Microsoft JhengHei UI", 9)).place(x=790, y=690)

        try:
            if getattr(self.master, "last_local_path", "") and not self.selected_file.get():
                self.selected_file.set(self.master.last_local_path)
        except Exception:
            pass

    def choose(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.selected_file.set(p)


# =========================================================
# PageC：151 比對（占位；保留你原本版本即可）
# （此段你原本就很長，且你目前需求是 PageD/PageE 先跑起來，因此保持原樣即可）
# =========================================================
class PageC(tk.Frame):
    def __init__(self, master, go_home):
        super().__init__(master, bg=BG_CARD)
        self.go_home = go_home
        self.file_local = tk.StringVar()
        self.file_qp111 = tk.StringVar()
        self.file_over = tk.StringVar()  # 可選：排班超過核定/額定數量表（C=姓名, H=超過數量）
        self.last_no_need_allocation = False
        self.last_output_path = ""
        self.build()

    def build(self):
        tk.Button(
            self, text="← 回首頁", command=self.go_home,
            bg=BG_CARD, fg=FG_MAIN, relief="flat", cursor="hand2"
        ).place(x=20, y=15)

        tk.Label(
            self, text="仁寶當月服務內容 VS 151 額分比對（需補分配）",
            font=("Microsoft JhengHei UI", 22, "bold"),
            bg=BG_CARD, fg=FG_MAIN
        ).place(relx=0.5, y=35, anchor="center")

        tk.Button(
            self, text="選擇【當月名單】",
            command=self.sel_local, bg="#4B5563", fg="white", cursor="hand2"
        ).place(x=60, y=95)
        tk.Label(self, textvariable=self.file_local, fg=FG_LINK, bg=BG_CARD).place(x=360, y=98)

        tk.Button(
            self, text="選擇【151額分產出名單】",
            command=self.sel_qp111, bg="#4B5563", fg="white", cursor="hand2"
        ).place(x=60, y=130)
        tk.Label(self, textvariable=self.file_qp111, fg=FG_LINK, bg=BG_CARD).place(x=360, y=133)

        tk.Button(
            self, text="選擇【檢查是否超過仁寶當月額度 Excel（可選）】",
            command=self.sel_over, bg="#4B5563", fg="white", cursor="hand2"
        ).place(x=60, y=165)
        tk.Label(self, textvariable=self.file_over, fg=FG_LINK, bg=BG_CARD).place(x=360, y=168)

        tk.Button(
            self, text="開始比對並輸出『需補分配』",
            width=24, height=2,
            command=lambda: threading.Thread(target=self.run, daemon=True).start(),
            bg="#3949AB", fg=FG_MAIN, activebackground="#303F9F",
            cursor="hand2"
        ).place(x=60, y=200)

        self.log = scrolledtext.ScrolledText(
            self, width=96, height=22,
            bg=BG_TEXT, fg=FG_MAIN,
            insertbackground=FG_MAIN, relief="flat"
        )
        self.log.place(x=60, y=270)

        tk.Label(
            self, text="FOR M. BY BEN",
            bg=self["bg"], fg="#666666",
            font=("Microsoft JhengHei UI", 10)
        ).place(relx=1.0, rely=1.0, x=-20, y=-20, anchor="se")

        # 若 PageB 剛存完，可自動帶入
        try:
            if getattr(self.master, "last_qp111_path", ""):
                self.file_qp111.set(self.master.last_qp111_path)
        except Exception:
            pass

    def sel_local(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.file_local.set(p)

    def sel_qp111(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.file_qp111.set(p)

    def sel_over(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.file_over.set(p)

    def _log(self, s: str):
        self.log.insert(tk.END, s + "\n")
        self.log.see(tk.END)

    @staticmethod
    def _norm_name(s: str) -> str:
        return str(s).replace(" ", "").replace("　", "").strip()

    @staticmethod
    def _norm_fee_source(s: str) -> str:
        return str(s).replace(" ", "").replace("　", "").strip()

    @staticmethod
    def _pick_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
        cols = set(map(str, df.columns))
        for c in candidates:
            if c in cols:
                return c
        return None

    # =========================================================
    # 服務內容輸出（以 PageA 為準）：
    # - 純自費項目（補助=0）不顯示
    # - 補助+自費：只顯示補助次數
    # - 只有補助：顯示補助次數
    # - 不顯示「補助/自費」字樣
    # =========================================================
    @staticmethod
    def _build_local_items_subsidy_only(code_map: dict) -> str:
        parts = []
        for code in sorted(code_map.keys()):
            b = int(code_map[code].get("補助", 0))
            if b > 0:
                parts.append(f"{code}：{b}")
        return "、".join(parts)

    # =========================================================
    # 不足項目(不足量)：只針對 BA（151 分配是 BA）
    # =========================================================
    @staticmethod
    def _build_lack_items_with_shortage_ba_only(code_map: dict, alloc_map: dict) -> str:
        parts = []
        for code in sorted(code_map.keys()):
            if not str(code).startswith("BA"):
                continue
            actual_b = int(code_map[code].get("補助", 0))
            if actual_b <= 0:
                continue
            alloc_n = int(alloc_map.get(code, 0))
            if actual_b > alloc_n:
                parts.append(f"{code}:少{actual_b - alloc_n}")
        return "、".join(parts)

    @staticmethod
    def _read_over_quota_excel_by_com(path: str) -> dict:
        """
        用 Excel COM 讀值（避免 openpyxl style 問題）
        - C2 起：個案姓名
        - H2 起：排班超過核定/額定數量
        回傳 dict[norm_name] = "超過數量(字串)"
        """
        if not path:
            return {}

        try:
            import os
            import pythoncom
            import win32com.client as win32
        except Exception as e:
            raise RuntimeError(
                "你有選『排班超過仁寶當月額度』檔案，但環境沒有 pywin32。\n"
                "請先安裝：pip install pywin32\n"
                f"原始錯誤：{e}"
            )

        over_map = {}
        pythoncom.CoInitialize()
        excel = None
        wb = None

        try:
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(os.path.abspath(path), ReadOnly=True)
            ws = wb.Worksheets(1)

            used = ws.UsedRange
            last_row = used.Row + used.Rows.Count - 1
            if last_row < 2:
                return over_map

            blank_streak = 0

            # C=3, H=8
            for r in range(2, last_row + 1):
                name = ws.Cells(r, 3).Value  # C
                if name is None:
                    blank_streak += 1
                    if blank_streak >= 30:
                        break
                    continue

                name = str(name).strip()
                if not name:
                    blank_streak += 1
                    if blank_streak >= 30:
                        break
                    continue

                blank_streak = 0

                raw = ws.Cells(r, 8).Value   # H
                if raw is None:
                    continue
                raw_s = str(raw).strip()
                if not raw_s:
                    continue

                try:
                    f = float(raw_s)
                    val = str(int(f)) if f.is_integer() else str(f)
                except Exception:
                    val = raw_s

                key = PageC._norm_name(name)
                if key not in over_map:
                    over_map[key] = val
                else:
                    if val not in over_map[key].split("、"):
                        over_map[key] = over_map[key] + "、" + val

            return over_map

        finally:
            try:
                if wb:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel:
                    excel.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def _build_respite_need_map(self, respite_rows: list[dict] | None) -> dict[str, dict]:
        out: dict[str, dict] = {}
        if not respite_rows:
            return out

        for r in respite_rows:
            person = str(r.get("個案姓名", "")).strip()
            if not person:
                continue
            key = self._norm_name(person)

            alloc = _safe_int(r.get("152_GA09(區間)"), default=0)
            actual = _safe_int(r.get("QD120實際_GA09(區間)"), default=0)
            diff = _safe_int(r.get("差額(應有-目前)"), default=0)
            note = str(r.get("備註", "")).strip()

            has_no_alloc = (alloc <= 0 and actual > 0)
            has_org_gap = (
                actual <= 0
                and ("機構篩選" in note)
                and bool(re.search(r"QD120A?:\d+", note))
            )
            has_no_range = ("查無資料或區間不包含目標月" in note)

            if diff <= 0 and (not has_no_alloc) and (not has_org_gap) and (not has_no_range):
                continue

            if diff > 0:
                text = f"GA09需補{diff}"
            elif has_no_alloc:
                text = f"GA09未分配(實際{actual})"
            elif has_no_range:
                text = "GA09查無資料/區間不含目標月"
            else:
                text = "GA09疑似缺單位/機構不符"

            if key not in out:
                out[key] = {"name": person, "texts": [text]}
            else:
                if text not in out[key]["texts"]:
                    out[key]["texts"].append(text)

        for k in list(out.keys()):
            out[k]["text"] = "；".join(out[k]["texts"])
        return out

    def run(self, save_path: str | None = None, auto_mode: bool = False,
            respite_rows: list[dict] | None = None, log_hook=None, progress_hook=None):
        self.log.delete(1.0, tk.END)
        self.last_no_need_allocation = False
        self.last_output_path = ""

        def emit(msg: str):
            self._log(msg)
            if callable(log_hook):
                try:
                    log_hook(msg)
                except Exception:
                    pass

        if not (self.file_local.get() and self.file_qp111.get()):
            messagebox.showerror("錯誤", "請先選兩份 Excel")
            return False

        # =========================================================
        # 0) 讀「排班超過核定數量」（可選）
        # =========================================================
        over_map = {}
        if self.file_over.get():
            try:
                over_map = self._read_over_quota_excel_by_com(self.file_over.get())
                emit(f"✅ 排班超過仁寶當月額度表讀取完成：{len(over_map)} 位")
            except Exception as e:
                messagebox.showerror("錯誤", f"讀取『排班超過仁寶當月額度』檔案失敗：\n{e}")
                return False
        else:
            emit("ℹ️ 未選擇『檢查是否超過仁寶當月額度』檔案：輸出欄位將留空")
        if callable(progress_hook):
            progress_hook(10)

        # =========================================================
        # 1) 讀 PageA（仁寶彙整結果）
        # =========================================================
        try:
            df_local = pd.read_excel(self.file_local.get(), dtype=str).fillna("")
            df_local.columns = ["月份", "姓名", "費用來源", "服務內容", "A單位名稱", "A個管姓名"]
        except Exception as e:
            messagebox.showerror("錯誤", f"PageA Excel 格式不符（需為 PageA 產出欄位）\n{e}")
            return False

        # ✅ PageA 服務內容解析：
        # 支援半形/全形括號 ( ) / （ ）
        # token-by-token 解析，並先跳過「純自費 token」
        token_pat = re.compile(
            r"([A-Z]{2}[0-9A-Za-z\-]+)\s*[:：]\s*(?:"
            r"(\d+)\s*[（(]補助[）)]\s*(?:/\s*(\d+)\s*[（(]自費[）)])?"
            r"|(\d+)\s*[（(]自費[）)]"
            r"|(\d+)"
            r")"
        )

        local = {}
        unit_map = {}
        mgr_map = {}
        fee_map = {}
        raw_name_map = {}

        for _, row in df_local.iterrows():
            raw_name = str(row["姓名"]).strip()
            key = self._norm_name(raw_name)

            fee_map[key] = self._norm_fee_source(row["費用來源"])
            unit_map[key] = str(row["A單位名稱"]).strip()
            mgr_map[key] = str(row["A個管姓名"]).strip()
            raw_name_map[key] = raw_name

            text = str(row["服務內容"]).strip()
            if key not in local:
                local[key] = {}

            # ✅ 先用頓號/逗號切 token，避免純自費被「(\d+)」分支誤判成補助
            tokens = re.split(r"[、,，]\s*", text)
            for tok in tokens:
                tok = (tok or "").strip()
                if not tok:
                    continue

                # ✅ 純自費 token（有自費、無補助、也沒有 /）直接跳過
                if ("自費" in tok) and ("補助" not in tok) and ("/" not in tok):
                    continue

                m = token_pat.search(tok)
                if not m:
                    continue

                code, sub, selfpay, only_self, noflag = m.groups()
                code = str(code).upper()  # ✅ 統一大寫，避免 BA17e/BA17E 對不上

                if code not in local[key]:
                    local[key][code] = {"補助": 0, "自費": 0}

                if sub:
                    local[key][code]["補助"] += int(sub)
                    if selfpay:
                        local[key][code]["自費"] += int(selfpay)
                elif only_self:
                    # 這裡理論上不會進來（純自費已被跳過），保險仍保留
                    local[key][code]["自費"] += int(only_self)
                elif noflag:
                    # 無標註視為補助
                    local[key][code]["補助"] += int(noflag)

        emit(f"✅ 當月名單(PageA) 讀取完成：{len(local)} 位案主")
        if callable(progress_hook):
            progress_hook(35)

        # =========================================================
        # 2) 讀 PageB（151 額分 QP111 結果）
        # =========================================================
        try:
            df_qp = pd.read_excel(self.file_qp111.get(), dtype=str).fillna("")
        except Exception as e:
            messagebox.showerror("錯誤", f"PageB Excel 讀取失敗\n{e}")
            return False

        col_name = self._pick_col(df_qp, ["姓名", "個案姓名"])
        col_ym   = self._pick_col(df_qp, ["分配年月", "yyyymm", "年月"])
        col_ba   = self._pick_col(df_qp, ["額分(BA碼:次數)", "額分", "qd100s", "qd100"])

        if not (col_name and col_ba):
            messagebox.showerror("錯誤", "PageB Excel 欄位不符：至少需要『姓名』與『額分(BA碼:次數)』")
            return False

        qp_map = {}
        qp_meta = {}

        for _, r in df_qp.iterrows():
            raw_name = str(r.get(col_name, "")).strip()
            if not raw_name:
                continue
            key = self._norm_name(raw_name)

            ym = str(r.get(col_ym, "")).strip() if col_ym else ""
            alloc_text = str(r.get(col_ba, "")).strip()

            found = re.findall(
                r"(BA[0-9A-Za-z\-]+)\s*[:：]\s*(\d+)",
                alloc_text,
                flags=re.IGNORECASE
            )
            alloc = {str(c).upper(): int(n) for c, n in found}

            if key not in qp_map:
                qp_map[key] = {}
            for c, n in alloc.items():
                qp_map[key][c] = qp_map[key].get(c, 0) + int(n)

            if key not in qp_meta:
                qp_meta[key] = {"分配年月": ym}
            else:
                if ym:
                    qp_meta[key]["分配年月"] = ym

        emit(f"✅ 151 額分 讀取完成：{len(qp_map)} 位案主")
        if callable(progress_hook):
            progress_hook(55)

        # =========================================================
        # 3) 判定「需補分配」
        #    ✅ 若費用來源=自費 → 不比對、不呈現
        #    ✅ 只比較 BA 的補助次數 vs 151 分配
        # =========================================================
        flagged = set()
        skipped_selfpay_cases = 0

        for key, codes in local.items():
            if fee_map.get(key, "") == "自費":
                skipped_selfpay_cases += 1
                continue

            alloc = qp_map.get(key, {})
            for code, v in codes.items():
                if not str(code).startswith("BA"):
                    continue
                actual_b = int(v.get("補助", 0))
                if actual_b <= 0:
                    continue
                alloc_n = int(alloc.get(code, 0))
                if actual_b > alloc_n:
                    flagged.add(key)
                    break

        emit(f"ℹ️ 費用來源=自費，已排除：{skipped_selfpay_cases} 位")

        respite_need_map = self._build_respite_need_map(respite_rows)
        if respite_need_map:
            emit(f"✅ 已合併喘息補分配需求：{len(respite_need_map)} 位")

        combined_keys = set(flagged) | set(respite_need_map.keys())

        if not combined_keys:
            emit("✔ 無差異：PageA 的 BA 次數皆未超過 151 分配")
            self.last_no_need_allocation = True
            messagebox.showinfo("完成", "✔ 全部一致，無需分配！")
            if callable(progress_hook):
                progress_hook(100)
            return True

        # =========================================================
        # 4) 輸出 flagged（需補分配者）
        # - 彙總「服務內容」：PageA 全部碼別，但排除純自費，且混合只顯示補助數字
        # - 明細：只列 BA（151 只對 BA）
        # =========================================================
        summary_rows = []
        detail_rows = []

        for key in sorted(combined_keys):
            # 保險：再次排除自費
            if fee_map.get(key, "") == "自費" and key not in respite_need_map:
                continue

            codes = local.get(key, {})
            alloc = qp_map.get(key, {})
            ym = (qp_meta.get(key, {}) or {}).get("分配年月", "")
            over_val = over_map.get(key, "")
            respite_text = (respite_need_map.get(key, {}) or {}).get("text", "")

            service_text = self._build_local_items_subsidy_only(codes)
            lack_text = self._build_lack_items_with_shortage_ba_only(codes, alloc)

            person_name = raw_name_map.get(key, "")
            if not person_name:
                person_name = (respite_need_map.get(key, {}) or {}).get("name", key)

            summary_rows.append({
                "姓名": person_name,
                "服務內容(PageA-排除自費)": service_text,
                "A單位名稱": unit_map.get(key, ""),
                "A個管姓名": mgr_map.get(key, ""),
                "分配年月": ym,
                "費用來源": fee_map.get(key, ""),
                "不足項目(不足量)": lack_text,
                "排班超過仁寶當月額度": over_val,
                "喘息需分配": respite_text,
            })

            for code in sorted(codes.keys()):
                if not str(code).startswith("BA"):
                    continue
                actual_b = int(codes[code].get("補助", 0))
                if actual_b <= 0:
                    continue
                alloc_n = int(alloc.get(code, 0))
                is_lack = "是" if actual_b > alloc_n else "否"

                detail_rows.append({
                    "姓名": raw_name_map.get(key, key),
                    "B碼別": code,
                    "仁寶當月次數": actual_b,
                    "151_目前分配": alloc_n,
                    "是否不足": is_lack,
                    "A單位名稱": unit_map.get(key, ""),
                    "A個管姓名": mgr_map.get(key, ""),
                    "分配年月": ym,
                    "費用來源": fee_map.get(key, ""),
                    "排班超過仁寶當月額度": over_val,
                })

        df_sum = pd.DataFrame(summary_rows)
        df_detail = pd.DataFrame(detail_rows)

        # =========================================================
        # 5) 欄位順序：分配年月、費用來源放在個管姓名後面
        # =========================================================
        sum_cols = [
            "姓名",
            "服務內容(PageA-排除自費)",
            "A單位名稱",
            "A個管姓名",
            "分配年月",
            "費用來源",
            "不足項目(不足量)",
            "排班超過仁寶當月額度",
            "喘息需分配",
        ]
        detail_cols = [
            "姓名",
            "B碼別",
            "仁寶當月次數",
            "151_目前分配",
            "是否不足",
            "A單位名稱",
            "A個管姓名",
            "分配年月",
            "費用來源",
            "排班超過仁寶當月額度",
        ]

        df_sum = df_sum.reindex(columns=[c for c in sum_cols if c in df_sum.columns])
        df_detail = df_detail.reindex(columns=[c for c in detail_cols if c in df_detail.columns])

        emit(f"✅ 需補分配案主：{len(df_sum)} 位（明細 BA 筆數：{len(df_detail)} 筆）")
        if callable(progress_hook):
            progress_hook(80)

        # =========================================================
        # 6) 儲存 Excel（兩張分頁）
        # =========================================================
        save = save_path
        if not save:
            save = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile="151需補分配.xlsx",
                filetypes=[("Excel 檔案", "*.xlsx")]
            )
        if not save:
            if not auto_mode:
                messagebox.showwarning("未存檔", "你取消了存檔")
            return False

        try:
            with pd.ExcelWriter(save, engine="openpyxl") as writer:
                df_sum.to_excel(writer, index=False, sheet_name="彙總(需補分配)")
                df_detail.to_excel(writer, index=False, sheet_name="明細(BA比對)")
        except Exception as e:
            messagebox.showerror("錯誤", f"輸出失敗：\n{e}")
            return False

        emit(f"🎉 已輸出：{save}")
        self.last_output_path = save
        if not auto_mode:
            messagebox.showinfo("完成", f"已輸出：\n{save}")
        if callable(progress_hook):
            progress_hook(100)
        return True


# =========================================================
# PageD：152（QP211 區間）＋ QD120/QD120A 實際 GA09
# =========================================================
class PageD(tk.Frame):
    def __init__(self, master, go_home):
        super().__init__(master, bg=BG_CARD)
        self.go_home = go_home

        self.username = tk.StringVar()
        self.password = tk.StringVar()
        self.selected_file = tk.StringVar()   # GA09 名單（PageA 產出）
        self.target_ym = tk.StringVar()
        self.progress = tk.IntVar()

        self.org_var = tk.StringVar(value=ORG_OPTIONS[0])

        self.session = None
        self.results = []

        self.build()

        try:
            if getattr(self.master, "last_target_ym", ""):
                self.target_ym.set(self.master.last_target_ym)
        except Exception:
            pass

        try:
            if getattr(self.master, "last_ga09_path", ""):
                self.selected_file.set(self.master.last_ga09_path)
        except Exception:
            pass

    def _ui_log(self, msg: str):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)

    def append_log(self, msg: str):
        self.after(0, lambda: self._ui_log(msg))

    def update_progress(self, now, total):
        def _upd():
            self.progress.set(0 if total <= 0 else int(now / total * 100))
            self.bar.update()
        self.after(0, _upd)

    def tree_clear(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

    def tree_add_row(self, row: dict):
        diff_val = 0
        try:
            diff_val = int(str(row.get("差額(應有-目前)", "0")).strip() or 0)
        except Exception:
            diff_val = 0

        note_text = str(row.get("備註", "")).strip()
        is_no_range = ("查無資料或區間不包含目標月" in note_text)

        tags = ("insufficient",) if (diff_val > 0 or is_no_range) else ()
        self.tree.insert(
            "", "end",
            values=(
                row.get("個案姓名", ""),
                row.get("目標年月", ""),
                row.get("152_區間(yyymm)", ""),
                row.get("152_GA09(區間)", ""),
                row.get("QD120實際_GA09(區間)", ""),
                row.get("152應有_GA09", ""),
                row.get("差額(應有-目前)", ""),
                row.get("case_ids", ""),
                row.get("152_lastUpdated", ""),
                row.get("備註", "")
            ),
            tags=tags
        )

    def _get_or_login_session(self) -> requests.Session | None:
        if getattr(self.master, "shared_session", None):
            self.session = self.master.shared_session
            return self.session
        if self.session:
            return self.session

        user = self.username.get().strip() or getattr(self.master, "shared_username", tk.StringVar()).get().strip()
        pwd = self.password.get().strip() or getattr(self.master, "shared_password", tk.StringVar()).get().strip()
        if not user or not pwd:
            messagebox.showerror("錯誤", "請先輸入 LCMS 帳號與密碼")
            return None

        self.username.set(user)
        self.password.set(pwd)
        if hasattr(self.master, "shared_username"):
            self.master.shared_username.set(user)
        if hasattr(self.master, "shared_password"):
            self.master.shared_password.set(pwd)

        self.append_log("🔐 正在登入 LCMS（驗證碼/Cloudflare/reCAPTCHA 需人工處理）...")
        sess = lcms_login_gui(self.master, user, pwd)
        if not sess:
            return None
        self.session = sess
        self.master.shared_session = sess
        return sess

    @staticmethod
    def _norm_name(s: str) -> str:
        return str(s).replace(" ", "").replace("　", "").strip()

    def run_query(self, save_path: str | None = None, auto_mode: bool = False,
                  org_selected_override: str | None = None, log_hook=None, progress_hook=None):
        self.btn_start.config(state="disabled")
        self.log.delete(1.0, tk.END)
        self.tree_clear()
        self.results = []

        def emit(msg: str):
            self.append_log(msg)
            if callable(log_hook):
                try:
                    log_hook(msg)
                except Exception:
                    pass

        target = self.target_ym.get().strip()
        if len(target) != 5 or (not target.isdigit()):
            messagebox.showerror("錯誤", "請輸入正確的『目標年月』，例如：11501、11503")
            self.btn_start.config(state="normal")
            return False

        fp = self.selected_file.get()
        if not fp:
            messagebox.showerror("錯誤", "請先選擇【GA09 喘息名單")
            self.btn_start.config(state="normal")
            return False

        org_selected = (org_selected_override or self.org_var.get().strip() or ORG_NO_FILTER).strip()
        if org_selected_override:
            self.org_var.set(org_selected)

        # 1) 讀 GA09 名單（只拿姓名當母體）
        try:
            df_ga09 = pd.read_excel(fp, dtype=str).fillna("")
        except Exception as e:
            messagebox.showerror("錯誤", f"GA09 名單讀取失敗\n{e}")
            self.btn_start.config(state="normal")
            return False

        if "個案姓名" not in df_ga09.columns:
            try:
                raw = pd.read_excel(fp, header=None)
                tmp = raw.iloc[1:, [1]].copy()  # 只取 B 欄
                tmp.columns = ["個案姓名"]
                df_ga09 = tmp.fillna("")
            except Exception as e:
                messagebox.showerror("錯誤", f"GA09 名單欄位不符且保底讀取失敗\n{e}")
                self.btn_start.config(state="normal")
                return False

        raw_name_map = {}
        for _, r in df_ga09.iterrows():
            nm = str(r.get("個案姓名", "")).strip()
            if not nm:
                continue
            key = self._norm_name(nm)
            raw_name_map[key] = nm

        keys = sorted(raw_name_map.keys())
        if not keys:
            messagebox.showwarning("提示", "GA09 名單內沒有任何姓名")
            self.btn_start.config(state="normal")
            return False

        emit(f"🟢 名單共 {len(keys)} 位（以姓名為母體）")
        emit(f"🟢 目標年月：{target}")
        emit(f"🟢 長照機構篩選：{org_selected}")
        emit("🧠 規則：先找 152 區間，再用『民國日期 QD120/QD120A』計算『實際 GA09 使用量』")
        emit("    若 實際 > 152顯示 → 152應有=實際（不足）\n")

        # 2) 登入
        sess = self._get_or_login_session()
        if not sess:
            emit("❌ 登入失敗或取消。")
            self.btn_start.config(state="normal")
            return False
        try:
            setattr(sess, "_retry_log_hook", emit)
        except Exception:
            pass

        total = len(keys)

        for idx, key in enumerate(keys, start=1):
            person = raw_name_map.get(key, key)
            emit(f"[{idx}/{total}] {person}")

            try:
                matched_rows = api_qp211_query_person_matched_rows(sess, person, target)
                alloc_ga09, ranges_list, lu, note = summarize_qp211_matched(matched_rows)

                if alloc_ga09 is None or not ranges_list:
                    row_out = {
                        "個案姓名": person,
                        "目標年月": target,
                        "152_區間(yyymm)": "",
                        "152_GA09(區間)": "",
                        "QD120實際_GA09(區間)": "",
                        "152應有_GA09": "",
                        "差額(應有-目前)": "",
                        "case_ids": "",
                        "152_lastUpdated": "",
                        "備註": note or "查無 152 區間"
                    }
                    self.results.append(row_out)
                    self.after(0, lambda r=row_out: self.tree_add_row(r))
                    emit(f"   ❌ 152 查無區間：{row_out['備註']}")
                    self.update_progress(idx, total)
                    if callable(progress_hook):
                        try:
                            progress_hook(int(idx * 100 / total))
                        except Exception:
                            pass
                    continue

                actual_sum = 0
                case_ids_union: list[str] = []
                actual_notes: list[str] = []

                for rng in ranges_list:
                    pr = parse_yyymm_range(rng)
                    if not pr:
                        actual_notes.append(f"區間格式無法解析：{rng}")
                        continue
                    roc_start, roc_end = pr

                    # ✅ 這裡原本缺函式，現在已補齊
                    actual_ga09, case_ids, n = calc_actual_ga09_by_name_and_range(
                        sess, person, roc_start, roc_end, org_filter=org_selected
                    )
                    actual_sum += int(actual_ga09)
                    for cid in case_ids:
                        if cid not in case_ids_union:
                            case_ids_union.append(cid)
                    if n:
                        actual_notes.append(f"{rng}:{n}")

                alloc = int(alloc_ga09 or 0)
                actual = int(actual_sum)

                should_be = max(alloc, actual)
                diff = should_be - alloc

                if actual > alloc:
                    note2 = f"不足：實際{actual} > 152顯示{alloc} → 152應有={should_be}"
                else:
                    note2 = f"足夠：實際{actual} ≤ 152顯示{alloc} → 不需調整"

                if actual_notes:
                    note2 = note2 + "；" + "；".join(actual_notes)

                row_out = {
                    "個案姓名": person,
                    "目標年月": target,
                    "152_區間(yyymm)": "、".join(ranges_list),
                    "152_GA09(區間)": alloc,
                    "QD120實際_GA09(區間)": actual,
                    "152應有_GA09": should_be,
                    "差額(應有-目前)": diff if diff > 0 else 0,
                    "case_ids": "、".join(case_ids_union),
                    "152_lastUpdated": lu,
                    "備註": note2
                }
                self.results.append(row_out)
                self.after(0, lambda r=row_out: self.tree_add_row(r))
                emit(f"   ✅ 區間：{'、'.join(ranges_list)}；152={alloc}；實際={actual}；應有={should_be}")

            except Exception as e:
                row_out = {
                    "個案姓名": person,
                    "目標年月": target,
                    "152_區間(yyymm)": "",
                    "152_GA09(區間)": "",
                    "QD120實際_GA09(區間)": "",
                    "152應有_GA09": "",
                    "差額(應有-目前)": "",
                    "case_ids": "",
                    "152_lastUpdated": "",
                    "備註": f"異常：{e}"
                }
                self.results.append(row_out)
                self.after(0, lambda r=row_out: self.tree_add_row(r))
                emit(f"   ❌ 異常：{e}")

            self.update_progress(idx, total)
            if callable(progress_hook):
                try:
                    progress_hook(int(idx * 100 / total))
                except Exception:
                    pass

        if not save_path:
            save_path = filedialog.asksaveasfilename(
                title="另存 152 區間＋實際喘息校對結果",
                defaultextension=".xlsx",
                initialfile=f"{target}_152_區間vs實際(GA09).xlsx",
                filetypes=[("Excel 檔案", "*.xlsx")]
            )
        if save_path:
            try:
                pd.DataFrame(self.results).to_excel(save_path, index=False)

                # ✅ 將「不足個案」整列標紅
                try:
                    wb = load_workbook(save_path)
                    ws = wb.active
                    headers = [c.value for c in ws[1]]
                    diff_idx = None
                    for i, h in enumerate(headers, start=1):
                        if str(h).strip() == "差額(應有-目前)":
                            diff_idx = i
                            break

                    if diff_idx:
                        red_font = Font(color="FF4D4F")
                        for r in range(2, ws.max_row + 1):
                            val = ws.cell(row=r, column=diff_idx).value
                            try:
                                if int(str(val).strip() or 0) > 0:
                                    for c in range(1, ws.max_column + 1):
                                        ws.cell(row=r, column=c).font = red_font
                            except Exception:
                                continue
                    wb.save(save_path)
                except Exception:
                    pass

                self.master.last_qp211_path = save_path
                self.master.last_target_ym = target
                self.master.last_ga09_path = fp
                if not auto_mode:
                    messagebox.showinfo("完成", f"已儲存：\n{save_path}")
                emit(f"\n🎉 完成！已輸出：{save_path}")
            except Exception as e:
                messagebox.showerror("錯誤", f"儲存 Excel 失敗：\n{e}")
                self.btn_start.config(state="normal")
                return False
        else:
            if not auto_mode:
                messagebox.showwarning("未儲存", "你取消了存檔（畫面仍保留查詢結果）")
            self.btn_start.config(state="normal")
            return False

        self.btn_start.config(state="normal")
        if callable(progress_hook):
            try:
                progress_hook(100)
            except Exception:
                pass
        return True

    def build(self):
        tk.Button(
            self, text="← 回首頁", command=self.go_home,
            bg=BG_CARD, fg=FG_MAIN, relief="flat", cursor="hand2"
        ).place(x=20, y=15)

        tk.Label(
            self, text="152 區間＋ 實際喘息校對（GA09）",
            font=("Microsoft JhengHei UI", 18, "bold"),
            bg=BG_CARD, fg=FG_MAIN
        ).place(relx=0.5, y=20, anchor="center")

        tk.Label(self, text="LCMS 帳號：", bg=BG_CARD, fg=FG_MAIN).place(x=60, y=80)
        tk.Entry(self, textvariable=self.username, width=18,
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=150, y=80)

        tk.Label(self, text="LCMS 密碼：", bg=BG_CARD, fg=FG_MAIN).place(x=340, y=80)
        tk.Entry(self, textvariable=self.password, width=18, show="*",
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=430, y=80)

        tk.Button(
            self, text="選擇【GA09 名單】",
            command=self.choose, width=30,
            bg="#4B5563", fg="white", cursor="hand2"
        ).place(x=60, y=115)
        tk.Label(self, textvariable=self.selected_file, fg=FG_LINK, bg=BG_CARD).place(x=340, y=118)

        tk.Label(self, text="目標年月：", bg=BG_CARD, fg=FG_MAIN).place(x=60, y=155)
        tk.Entry(self, textvariable=self.target_ym, width=10,
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=150, y=155)
        tk.Label(self, text="例：11501", bg=BG_CARD, fg=FG_HINT).place(x=245, y=155)

        tk.Label(self, text="長照機構：", bg=BG_CARD, fg=FG_MAIN).place(x=330, y=155)
        cmb = ttk.Combobox(self, textvariable=self.org_var, values=ORG_OPTIONS, state="readonly", width=44)
        cmb.place(x=410, y=155)

        self.btn_start = tk.Button(
            self, text="開始校對", width=14, height=1,
            command=lambda: threading.Thread(target=self.run_query, daemon=True).start(),
            bg=ACCENT_OK, fg="white",
            font=("Microsoft JhengHei UI", 12, "bold"),
            cursor="hand2"
        )
        self.btn_start.place(x=730, y=140)

        self.bar = ttk.Progressbar(self, variable=self.progress, length=820)
        self.bar.place(x=60, y=190)

        cols = (
            "個案姓名", "目標年月",
            "152_區間(yyymm)", "152_GA09(區間)",
            "QD120實際_GA09(區間)", "152應有_GA09", "差額(應有-目前)",
            "case_ids", "152_lastUpdated", "備註"
        )
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=9)
        for c in cols:
            self.tree.heading(c, text=c)

        self.tree.tag_configure("insufficient", foreground="#FF4D4F")

        self.tree.column("個案姓名", width=110, anchor="w")
        self.tree.column("目標年月", width=80, anchor="center")
        self.tree.column("152_區間(yyymm)", width=140, anchor="w")
        self.tree.column("152_GA09(區間)", width=95, anchor="center")
        self.tree.column("QD120實際_GA09(區間)", width=125, anchor="center")
        self.tree.column("152應有_GA09", width=95, anchor="center")
        self.tree.column("差額(應有-目前)", width=105, anchor="center")
        self.tree.column("case_ids", width=140, anchor="w")
        self.tree.column("152_lastUpdated", width=105, anchor="center")
        self.tree.column("備註", width=260, anchor="w")

        self.tree.place(x=60, y=215, width=820, height=250)

        self.log = scrolledtext.ScrolledText(
            self, width=117, height=10,
            bg="#111111", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat"
        )
        self.log.place(x=60, y=480)

        tk.Label(self, text="FOR M. BY BEN", bg=BG_CARD, fg="#777777",
                 font=("Microsoft JhengHei UI", 9)).place(x=790, y=690)

    def choose(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.selected_file.set(p)


# =========================================================
# PageE：151 名單 vs QD120/QD120A（當月）比對 BA（只顯示有問題）
# =========================================================
class PageE(tk.Frame):
    def __init__(self, master, go_home):
        super().__init__(master, bg=BG_CARD)
        self.go_home = go_home

        self.username = tk.StringVar()
        self.password = tk.StringVar()
        self.selected_file = tk.StringVar()  # PageB 輸出（151額分）
        self.target_ym = tk.StringVar()
        self.progress = tk.IntVar()

        self.org_var = tk.StringVar(value=ORG_OPTIONS[0])

        self.session = None
        self.results = []

        self.build()

        try:
            if getattr(self.master, "last_qp111_path", ""):
                self.selected_file.set(self.master.last_qp111_path)
        except Exception:
            pass
        try:
            if getattr(self.master, "last_target_ym", ""):
                self.target_ym.set(self.master.last_target_ym)
        except Exception:
            pass

    def _ui_log(self, msg: str):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)

    def append_log(self, msg: str):
        self.after(0, lambda: self._ui_log(msg))

    def update_progress(self, now, total):
        def _upd():
            self.progress.set(0 if total <= 0 else int(now / total * 100))
            self.bar.update()
        self.after(0, _upd)

    def tree_clear(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

    def tree_add_row(self, row: dict):
        self.tree.insert(
            "", "end",
            values=(
                row.get("個案姓名", ""),
                row.get("分配年月", ""),
                row.get("機構(篩選)", ""),
                row.get("151_BA", ""),
                row.get("QD_BA(當月)", ""),
                row.get("是否有問題(151<QD)", ""),
                row.get("差異摘要", ""),
                row.get("case_ids", ""),
                row.get("備註", "")
            )
        )

    def _get_or_login_session(self) -> requests.Session | None:
        if getattr(self.master, "shared_session", None):
            self.session = self.master.shared_session
            return self.session
        if self.session:
            return self.session

        user = self.username.get().strip() or getattr(self.master, "shared_username", tk.StringVar()).get().strip()
        pwd = self.password.get().strip() or getattr(self.master, "shared_password", tk.StringVar()).get().strip()
        if not user or not pwd:
            messagebox.showerror("錯誤", "請先輸入 LCMS 帳號與密碼")
            return None

        self.username.set(user)
        self.password.set(pwd)
        if hasattr(self.master, "shared_username"):
            self.master.shared_username.set(user)
        if hasattr(self.master, "shared_password"):
            self.master.shared_password.set(pwd)

        self.append_log("🔐 正在登入 LCMS（驗證碼/Cloudflare/reCAPTCHA 需人工處理）...")
        sess = lcms_login_gui(self.master, user, pwd)
        if not sess:
            return None
        self.session = sess
        self.master.shared_session = sess
        return sess

    @staticmethod
    def _pick_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
        cols = set(map(str, df.columns))
        for c in candidates:
            if c in cols:
                return c
        return None

    def choose(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.selected_file.set(p)

    def run_query(self):
        self.btn_start.config(state="disabled")
        self.log.delete(1.0, tk.END)
        self.tree_clear()
        self.results = []

        ym = self.target_ym.get().strip()
        if len(ym) != 5 or (not ym.isdigit()):
            messagebox.showerror("錯誤", "請輸入正確的『分配年月』，例如：11412")
            self.btn_start.config(state="normal")
            return

        fp = self.selected_file.get()
        if not fp:
            messagebox.showerror("錯誤", "請先選擇【151 額分名單】")
            self.btn_start.config(state="normal")
            return

        org_selected = self.org_var.get().strip() or ORG_NO_FILTER

        try:
            df = pd.read_excel(fp, dtype=str).fillna("")
        except Exception as e:
            messagebox.showerror("錯誤", f"讀取 151 Excel 失敗\n{e}")
            self.btn_start.config(state="normal")
            return

        col_name = self._pick_col(df, ["姓名", "個案姓名"])
        col_ym   = self._pick_col(df, ["分配年月", "yyyymm", "年月"])
        col_ba   = self._pick_col(df, ["額分(BA碼:次數)", "額分", "qd100s", "qd100"])

        if not (col_name and col_ba):
            messagebox.showerror("錯誤", "151 Excel 欄位不符：至少需要『姓名』與『額分(BA碼:次數)』")
            self.btn_start.config(state="normal")
            return

        alloc_map_by_name: dict[str, dict] = {}
        raw_name_map: dict[str, str] = {}
        ym_by_name: dict[str, str] = {}

        for _, r in df.iterrows():
            person = str(r.get(col_name, "")).strip()
            if not person:
                continue
            raw_name_map[person] = person

            ym_row = str(r.get(col_ym, "")).strip() if col_ym else ym
            if ym_row:
                ym_by_name[person] = ym_row

            alloc_text = str(r.get(col_ba, "")).strip()
            found = re.findall(r"(BA[0-9A-Za-z\-]+)\s*[:：]\s*(\d+)", alloc_text, flags=re.IGNORECASE)
            alloc = {str(c).upper(): int(n) for c, n in found}

            if person not in alloc_map_by_name:
                alloc_map_by_name[person] = {}
            alloc_map_by_name[person] = merge_int_maps(alloc_map_by_name[person], alloc)

        persons = list(raw_name_map.keys())
        if not persons:
            messagebox.showwarning("提示", "151 名單內沒有任何姓名")
            self.btn_start.config(state="normal")
            return

        self.append_log(f"🟢 151 名單共 {len(persons)} 位")
        self.append_log(f"🟢 目標年月（民國）：{ym}")
        self.append_log(f"🟢 長照機構篩選：{org_selected}")
        self.append_log("🧠 新規則：151 >= QD → 沒問題(不呈現)；151 < QD → 有問題(呈現)\n")

        sess = self._get_or_login_session()
        if not sess:
            self.append_log("❌ 登入失敗或取消。")
            self.btn_start.config(state="normal")
            return
        try:
            setattr(sess, "_retry_log_hook", self.append_log)
        except Exception:
            pass

        roc_yyyymm = int(ym)

        total = len(persons)
        problem_count = 0

        for idx, person in enumerate(persons, start=1):
            self.append_log(f"[{idx}/{total}] {person}")

            alloc_ba = alloc_map_by_name.get(person, {}) or {}
            alloc_text = format_ba_map(alloc_ba)

            try:
                case_ids, note_ca = api_case_find_case_ids_strict(sess, person)
                if not case_ids:
                    row_out = {
                        "個案姓名": person,
                        "分配年月": ym_by_name.get(person, ym),
                        "機構(篩選)": org_selected,
                        "151_BA": alloc_text,
                        "QD_BA(當月)": "",
                        "是否有問題(151<QD)": "是",
                        "差異摘要": "查無 case_id（無法比對）",
                        "case_ids": "",
                        "備註": note_ca or "CA 查無 case_id"
                    }
                    self.results.append(row_out)
                    self.after(0, lambda r=row_out: self.tree_add_row(r))
                    problem_count += 1
                    self.append_log(f"   ⚠ 有問題：{row_out['差異摘要']}")
                    self.update_progress(idx, total)
                    continue

                used_union: dict[str, int] = {}
                qd_notes = []
                for cid in case_ids:
                    used_map, qd_note = calc_actual_ba_by_case_and_month(
                        sess, cid, roc_yyyymm, org_filter=org_selected
                    )
                    used_union = merge_int_maps(used_union, used_map)
                    if qd_note:
                        qd_notes.append(f"{cid}:{qd_note}")

                used_text = format_ba_map(used_union)

                has_problem, diff = compare_ba_alloc_vs_used(alloc_ba, used_union)

                if has_problem:
                    row_out = {
                        "個案姓名": person,
                        "分配年月": ym_by_name.get(person, ym),
                        "機構(篩選)": org_selected,
                        "151_BA": alloc_text,
                        "QD_BA(當月)": used_text,
                        "是否有問題(151<QD)": "是",
                        "差異摘要": diff,
                        "case_ids": "、".join(case_ids),
                        "備註": (note_ca + ("；" if note_ca and qd_notes else "") + "；".join(qd_notes)).strip("；")
                    }
                    self.results.append(row_out)
                    self.after(0, lambda r=row_out: self.tree_add_row(r))
                    problem_count += 1
                    self.append_log(f"   ⚠ 有問題：{diff}")
                else:
                    self.append_log("   ✅ 沒問題（151 ≥ QD）")

            except Exception as e:
                row_out = {
                    "個案姓名": person,
                    "分配年月": ym_by_name.get(person, ym),
                    "機構(篩選)": org_selected,
                    "151_BA": alloc_text,
                    "QD_BA(當月)": "",
                    "是否有問題(151<QD)": "是",
                    "差異摘要": f"異常：{e}",
                    "case_ids": "",
                    "備註": "查詢異常"
                }
                self.results.append(row_out)
                self.after(0, lambda r=row_out: self.tree_add_row(r))
                problem_count += 1
                self.append_log(f"   ❌ 異常（視為有問題）：{e}")

            self.update_progress(idx, total)

        self.append_log(f"\n📌 比對完成：有問題個案 {problem_count} / {total} 位\n")

        if not self.results:
            messagebox.showinfo("完成", "全部個案皆沒問題（151 ≥ QD），無需輸出問題清單。")
            self.append_log("✅ 全部沒問題，未輸出 Excel。")
            self.btn_start.config(state="normal")
            return

        save_path = filedialog.asksaveasfilename(
            title="另存 PageE 有問題清單",
            defaultextension=".xlsx",
            initialfile=f"{ym}_有問題清單(151小於QD).xlsx",
            filetypes=[("Excel 檔案", "*.xlsx")]
        )
        if save_path:
            try:
                df_problem = pd.DataFrame(self.results)
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    df_problem.to_excel(writer, index=False, sheet_name="有問題(151<QD)")

                self.master.last_pagee_path = save_path
                messagebox.showinfo("完成", f"已儲存：\n{save_path}")
                self.append_log(f"\n🎉 已輸出有問題清單：{save_path}")
            except Exception as e:
                messagebox.showerror("錯誤", f"儲存 Excel 失敗：\n{e}")
        else:
            messagebox.showwarning("未儲存", "你取消了存檔（畫面仍保留有問題結果）")

        self.btn_start.config(state="normal")

    def build(self):
        tk.Button(
            self, text="← 回首頁", command=self.go_home,
            bg=BG_CARD, fg=FG_MAIN, relief="flat", cursor="hand2"
        ).place(x=20, y=15)

        tk.Label(
            self, text="151 名單 VS 個案當月服務紀錄比對（BA）",
            font=("Microsoft JhengHei UI", 18, "bold"),
            bg=BG_CARD, fg=FG_MAIN
        ).place(relx=0.5, y=20, anchor="center")

        tk.Label(self, text="LCMS 帳號：", bg=BG_CARD, fg=FG_MAIN).place(x=60, y=80)
        tk.Entry(self, textvariable=self.username, width=18,
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=150, y=80)

        tk.Label(self, text="LCMS 密碼：", bg=BG_CARD, fg=FG_MAIN).place(x=340, y=80)
        tk.Entry(self, textvariable=self.password, width=18, show="*",
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=430, y=80)

        tk.Button(
            self, text="選擇【151 額分名單】",
            command=self.choose, width=32,
            bg="#4B5563", fg="white", cursor="hand2"
        ).place(x=60, y=115)
        tk.Label(self, textvariable=self.selected_file, fg=FG_LINK, bg=BG_CARD).place(x=340, y=118)

        tk.Label(self, text="分配年月：", bg=BG_CARD, fg=FG_MAIN).place(x=60, y=155)
        tk.Entry(self, textvariable=self.target_ym, width=10,
                 bg="#2D2D2D", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat").place(x=150, y=155)
        tk.Label(self, text="例：11412", bg=BG_CARD, fg=FG_HINT).place(x=245, y=155)

        tk.Label(self, text="長照機構：", bg=BG_CARD, fg=FG_MAIN).place(x=330, y=155)
        cmb = ttk.Combobox(self, textvariable=self.org_var, values=ORG_OPTIONS, state="readonly", width=44)
        cmb.place(x=410, y=155, height=24)

        self.btn_start = tk.Button(
            self, text="開始比對", width=12, height=1,
            command=lambda: threading.Thread(target=self.run_query, daemon=True).start(),
            bg=ACCENT_OK, fg="white",
            font=("Microsoft JhengHei UI", 12, "bold"),
            cursor="hand2"
        )
        self.btn_start.place(x=750, y=150)

        self.bar = ttk.Progressbar(self, variable=self.progress, length=820)
        self.bar.place(x=60, y=190)

        cols = (
            "個案姓名", "分配年月", "機構(篩選)", "151_BA", "QD_BA(當月)",
            "是否有問題(151<QD)", "差異摘要", "case_ids", "備註"
        )
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=9)
        for c in cols:
            self.tree.heading(c, text=c)

        self.tree.column("個案姓名", width=100, anchor="w")
        self.tree.column("分配年月", width=70, anchor="center")
        self.tree.column("機構(篩選)", width=170, anchor="w")
        self.tree.column("151_BA", width=150, anchor="w")
        self.tree.column("QD_BA(當月)", width=150, anchor="w")
        self.tree.column("是否有問題(151<QD)", width=110, anchor="center")
        self.tree.column("差異摘要", width=210, anchor="w")
        self.tree.column("case_ids", width=110, anchor="w")
        self.tree.column("備註", width=220, anchor="w")

        self.tree.place(x=60, y=215, width=820, height=250)

        self.log = scrolledtext.ScrolledText(
            self, width=117, height=10,
            bg="#111111", fg=FG_MAIN, insertbackground=FG_MAIN, relief="flat"
        )
        self.log.place(x=60, y=480)

        tk.Label(self, text="FOR M. BY BEN", bg=BG_CARD, fg="#777777",
                 font=("Microsoft JhengHei UI", 9)).place(x=790, y=690)


# =========================================================
# Advanced Page：進階功能入口（②③④）
# =========================================================
class AdvancedPage(tk.Frame):
    def __init__(self, master, go_home, goto_b, goto_c, goto_d):
        super().__init__(master, bg=BG_MAIN)
        self.go_home = go_home
        self.goto_b = goto_b
        self.goto_c = goto_c
        self.goto_d = goto_d
        self.build()

    def build(self):
        W, H = 900, 720
        card_w, card_h = 660, 84
        gap_y = 26

        canvas = tk.Canvas(self, width=W, height=H, bg=BG_MAIN, highlightthickness=0)
        canvas.pack(fill="both", expand=True)

        tk.Button(
            self, text="← 回首頁", command=self.go_home,
            bg=BG_MAIN, fg=FG_MAIN, relief="flat", cursor="hand2"
        ).place(x=20, y=15)

        canvas.create_text(W//2, 52, text="進階功能",
                           fill=FG_MAIN, font=("Microsoft JhengHei UI", 26, "bold"))

        def make_card(y, text, cb):
            x = (W - card_w)//2
            box = draw_round_rect(canvas, x, y, x+card_w, y+card_h, r=22,
                                  fill=BG_CARD, outline="#444444", width=2)
            txt = canvas.create_text(x+30, y+card_h/2, anchor="w",
                                     text=text, fill=FG_MAIN,
                                     font=("Microsoft JhengHei UI", 17, "bold"))

            def enter(_): canvas.itemconfig(box, fill="#2A2A2A")
            def leave(_): canvas.itemconfig(box, fill=BG_CARD)
            def click(_): cb()

            for item in (box, txt):
                canvas.tag_bind(item, "<Enter>", enter)
                canvas.tag_bind(item, "<Leave>", leave)
                canvas.tag_bind(item, "<Button-1>", click)

        y0 = 130
        make_card(y0,                          "151 額分查詢", self.goto_b)
        make_card(y0 + (card_h + gap_y)*1,     "額分校對（仁寶 VS 151，需補分配）", self.goto_c)
        make_card(y0 + (card_h + gap_y)*2,     "152 額分+實際喘息次數校對（GA09）", self.goto_d)

        canvas.create_text(W-20, H-23, anchor="se",
                           text="FOR M. BY BEN",
                           fill="#666666", font=("Microsoft JhengHei UI", 10))


# =========================================================
# App
# =========================================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("額分次數查詢系統")
        self.geometry("900x720")
        self.resizable(False, False)
        self.config(bg=BG_MAIN)

        setup_ttk_style(self)

        # 共用狀態
        self.shared_session = None
        self.shared_username = tk.StringVar()
        self.shared_password = tk.StringVar()
        self.last_local_path = ""
        self.last_ga09_path = ""
        self.last_qp111_path = ""
        self.last_qp211_path = ""
        self.last_list_path = ""
        self.last_target_ym = ""
        self.last_pagee_path = ""

        self.page_a = PageA(self, self.show_home)
        self.page_b = PageB(self, self.show_home)
        self.page_c = PageC(self, self.show_home)
        self.page_d = PageD(self, self.show_home)
        self.page_e = PageE(self, self.show_home)
        self.page_adv = AdvancedPage(self, self.show_home, self.show_b, self.show_c, self.show_d)

        self.home = HomePage(
            self, self.show_a, self.show_b, self.show_c, self.show_d, self.show_e, self.show_adv,
            self.clear_session, self.shared_username, self.shared_password
        )

        self.current = None
        self.show_home()

    def show(self, frame):
        if self.current:
            self.current.pack_forget()
        self.current = frame
        self.current.pack(fill="both", expand=True)

    def clear_session(self):
        self.shared_session = None
        self.shared_username.set("")
        self.shared_password.set("")
        for page in (self.page_a, self.page_b, self.page_c, self.page_d, self.page_e):
            if hasattr(page, "session"):
                try:
                    page.session = None
                except Exception:
                    pass
            if hasattr(page, "username"):
                try:
                    page.username.set("")
                except Exception:
                    pass
            if hasattr(page, "password"):
                try:
                    page.password.set("")
                except Exception:
                    pass
        messagebox.showinfo("完成", "已清除登入快取，請重新登入新帳號。")

    def show_home(self):
        self.show(self.home)

    def show_a(self):
        self.show(self.page_a)

    def show_b(self):
        self.show(self.page_b)

    def show_c(self):
        self.show(self.page_c)

    def show_d(self):
        self.show(self.page_d)

    def show_e(self):
        self.show(self.page_e)

    def show_adv(self):
        self.show(self.page_adv)


if __name__ == "__main__":
    App().mainloop()
