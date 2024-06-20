# -*- coding: utf-8 -*-
"""
Created on Thu Jun 20 21:31:13 2024

@author: USER
"""

import openpyxl
file="sales.xlsx"
wb=openpyxl.load_workbook(file,data_only=True)

#ws=wb.active
ws=wb['2024Q1']#可直接抓相對應的工作表
print("目前工作表:",ws.title)
print(ws['A4'].value)
print(ws['C2'].value)
print("總列數:",ws.max_row)
print("總欄數:",ws.max_column)
for i in range(3,ws.max_row+1):
    for x in range(1,ws.max_column+1):
     print(ws.cell(i,x).value,end=",")
print()